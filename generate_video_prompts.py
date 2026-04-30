"""
AI-Powered Video Generation Strategy & Prompt Creator
Multi-phase pipeline for intelligent video content generation with cultural accuracy
"""

import openpyxl
import json
import os
import time
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dotenv import load_dotenv
from google import genai


class VideoPromptGenerator:
    """
    Video generation system with intelligent strategy selection:
    - Analyzes song content to determine optimal video approach
    - Generates culturally accurate, 4K-ready prompts
    - Handles various video types: AI-generated, slideshows, simple animations
    - Splits long videos into Sora-compatible segments
    """
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.video_output_dir = self.excel_path.parent / "videoContent"
        self.song_content_dir = self.excel_path.parent / "content" / "FinalizedSongs"
        self.video_output_dir.mkdir(exist_ok=True)
        
        # Rate limiting settings - optimized for video generation
        self.phase_delay = 4  # Seconds between phases
        self.video_delay = 8  # Seconds between videos
        self.api_call_delay = 3  # Seconds between API calls within same video
        
        # Load Excel
        self.wb = openpyxl.load_workbook(excel_path)
        self.ws = self.wb.active
        
        # Parse columns
        self.headers = [cell.value for cell in self.ws[1]]
        self.column_map = {h: i for i, h in enumerate(self.headers)}
        print(f"📊 Video Excel Columns: {', '.join([h for h in self.headers if h])}")
        
        # Initialize Gemini
        self._init_gemini()
        
        print(f"✅ Video Generator Ready! Output: {self.video_output_dir}")
        print(f"⏱️  Rate Limits: {self.phase_delay}s phases, {self.video_delay}s videos, {self.api_call_delay}s API calls")
    
    def _init_gemini(self):
        """Initialize Gemini API"""
        load_dotenv()
        api_key = os.getenv('GEMINI_API_KEY')
        
        if not api_key or api_key == 'your_gemini_api_key_here':
            raise ValueError(
                "Set GEMINI_API_KEY in .env file\n"
                "Get key: https://makersuite.google.com/app/apikey"
            )
        
        self.client = genai.Client(api_key=api_key)
        self.model_name = os.getenv('GEMINI_MODEL', 'gemini-2.5-flash')
        print(f"✅ Gemini: {self.model_name}")
    
    def _get_cell_value(self, row: int, col_name: str) -> Optional[str]:
        """Get value from Excel cell"""
        if col_name not in self.column_map:
            return None
        col_idx = self.column_map[col_name]
        cell = self.ws.cell(row=row, column=col_idx + 1)
        return cell.value
    
    def _set_cell_value(self, row: int, col_name: str, value: str):
        """Set value in Excel cell"""
        if col_name in self.column_map:
            col_idx = self.column_map[col_name]
            self.ws.cell(row=row, column=col_idx + 1, value=value)
            self.wb.save(self.excel_path)
    
    def _extract_lyrics_from_song_file(self, script_filename: str) -> Dict[str, str]:
        """
        Extract relevant content from song file in FinalizedSongs folder.
        Returns dict with lyrics, title, description, etc.
        """
        song_path = self.song_content_dir / script_filename
        
        if not song_path.exists():
            raise FileNotFoundError(f"Song file not found: {song_path}")
        
        with open(song_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Extract sections
        result = {
            'full_content': content,
            'title': '',
            'lyrics': '',
            'description': '',
            'youtube_title': '',
            'youtube_description': ''
        }
        
        # Extract title
        title_match = re.search(r'SONG TITLE:\s*\n(.+?)(?:\n\n|\nHINDI LYRICS:)', content, re.DOTALL)
        if title_match:
            result['title'] = title_match.group(1).strip()
        
        # Extract lyrics (between HINDI LYRICS: and SUNO AI DESCRIPTION:)
        lyrics_match = re.search(r'HINDI LYRICS:\s*\n={80}\n(.+?)\n\nSUNO AI DESCRIPTION:', content, re.DOTALL)
        if lyrics_match:
            result['lyrics'] = lyrics_match.group(1).strip()
        
        # Extract Suno description
        desc_match = re.search(r'SUNO AI DESCRIPTION:\s*\n(.+?)(?:\n\nYOUTUBE TITLE:)', content, re.DOTALL)
        if desc_match:
            result['description'] = desc_match.group(1).strip()
        
        # Extract YouTube title
        yt_title_match = re.search(r'YOUTUBE TITLE:\s*\n(.+?)(?:\n\nYOUTUBE DESCRIPTION:)', content, re.DOTALL)
        if yt_title_match:
            result['youtube_title'] = yt_title_match.group(1).strip()
        
        # Extract YouTube description
        yt_desc_match = re.search(r'YOUTUBE DESCRIPTION:\s*\n(.+?)(?:\n\nTAGS:)', content, re.DOTALL)
        if yt_desc_match:
            result['youtube_description'] = yt_desc_match.group(1).strip()
        
        return result
    
    def _phase1_analyze_content(self, song_data: Dict, user_notes: Optional[str]) -> Dict:
        """
        Phase 1: Analyze song content to determine video strategy
        Returns: strategy type, complexity score, narrative depth, visual suggestions
        """
        print("  📖 Phase 1: Analyzing content...")
        
        prompt = f"""
You are a video content strategist for devotional music videos. Analyze this song and determine the optimal video production strategy.

SONG TITLE: {song_data['title']}
SONG DESCRIPTION: {song_data['description']}
LYRICS EXCERPT (first 500 chars):
{song_data['lyrics'][:500]}

USER NOTES: {user_notes or 'None'}

Analyze and return JSON with:
{{
  "content_type": "instrumental|prayer|narrative|epic_story",
  "narrative_complexity": 1-10 (1=simple prayer, 10=complex story),
  "visual_density": 1-10 (1=minimal visuals needed, 10=scene-heavy),
  "recommended_strategy": "simple_animated|slideshow|hybrid|full_ai",
  "cultural_elements": ["temple", "deity", "nature", etc],
  "key_scenes": ["scene1", "scene2", ...],
  "emotional_tone": "peaceful|devotional|energetic|epic|meditative",
  "ai_requirement_score": 1-10 (1=no AI needed, 10=heavy AI use),
  "reasoning": "brief explanation"
}}

STRATEGY GUIDELINES:
- instrumental/simple prayer (1-3 narrative) → simple_animated (background + audio visualizer)
- medium complexity (4-6 narrative) → slideshow or hybrid (images + some AI scenes)
- high complexity (7-10 narrative) → full_ai (AI-generated video scenes)
- Consider cultural sensitivity for deity representation
"""
        
        time.sleep(self.api_call_delay)
        response = self.client.models.generate_content(model=self.model_name, contents=prompt)
        
        try:
            # Extract JSON from response
            json_match = re.search(r'\{.*\}', response.text, re.DOTALL)
            if json_match:
                analysis = json.loads(json_match.group(0))
            else:
                raise ValueError("No JSON found in response")
            
            print(f"    ✅ Strategy: {analysis['recommended_strategy']} (AI score: {analysis['ai_requirement_score']}/10)")
            return analysis
        except Exception as e:
            print(f"    ⚠️ JSON parse error: {e}")
            # Fallback analysis
            return {
                "content_type": "prayer",
                "narrative_complexity": 5,
                "visual_density": 5,
                "recommended_strategy": "hybrid",
                "cultural_elements": [],
                "key_scenes": [],
                "emotional_tone": "devotional",
                "ai_requirement_score": 5,
                "reasoning": "Fallback analysis due to parsing error"
            }
    
    def _phase2_generate_video_structure(self, analysis: Dict, song_data: Dict, video_type: str) -> Dict:
        """
        Phase 2: Create video structure/timeline
        Returns: Timeline with scenes, durations, transition types
        """
        print("  🎬 Phase 2: Creating video structure...")
        
        # Implement 70% AI cap - even for full_ai strategy
        ai_score = analysis['ai_requirement_score']
        max_ai_percentage = 70  # Maximum 70% AI even for 10/10 AI score
        
        prompt = f"""
You are creating a video timeline structure for a devotional music video.

SONG: {song_data['title']}
STRATEGY: {analysis['recommended_strategy']}
CONTENT TYPE: {analysis['content_type']}
EMOTIONAL TONE: {analysis['emotional_tone']}
VIDEO TYPE: {video_type}
AI REQUIREMENT: {ai_score}/10 (but limit AI to ~{max_ai_percentage}% of total duration)

LYRICS (to base timing on):
{song_data['lyrics'][:1000]}

Create a JSON timeline structure:
{{
  "total_duration": "estimate in seconds",
  "segments": [
    {{
      "timestamp": "00:00-00:15",
      "duration_sec": 15,
      "segment_type": "intro|verse|chorus|bridge|outro|instrumental",
      "visual_approach": "static|slideshow|ai_scene|animation|audio_viz",
      "description": "what happens visually",
      "cultural_elements": ["specific elements to show"],
      "transition": "fade|cut|dissolve|zoom"
    }},
    ...
  ],
  "color_palette": ["#color1", "#color2", ...],
  "overall_style": "description"
}}

IMPORTANT:
- LIMIT ai_scene to maximum {max_ai_percentage}% of total video duration
- Use slideshow/static for simpler segments (transitions, instrumental breaks)
- Reserve AI for key story moments and climactic scenes only
- If video_type is "Long" or content is instrumental → use simple visuals (audio visualizer, deity image)
- Match segment timing to actual song structure from lyrics
- Each segment should be 10-30 seconds for variety
- Consider Sora's video length limits (break if needed)
"""
        
        time.sleep(self.api_call_delay)
        response = self.client.models.generate_content(model=self.model_name, contents=prompt)
        
        try:
            json_match = re.search(r'\{.*\}', response.text, re.DOTALL)
            if json_match:
                structure = json.loads(json_match.group(0))
            else:
                raise ValueError("No JSON in response")
            
            print(f"    ✅ Structure: {len(structure['segments'])} segments, ~{structure['total_duration']}s")
            return structure
        except Exception as e:
            print(f"    ⚠️ Structure parse error: {e}")
            return {
                "total_duration": "180",
                "segments": [{
                    "timestamp": "00:00-03:00",
                    "duration_sec": 180,
                    "segment_type": "full",
                    "visual_approach": "slideshow",
                    "description": "Devotional imagery",
                    "cultural_elements": [],
                    "transition": "fade"
                }],
                "color_palette": ["#FF6B35", "#F7931E"],
                "overall_style": "Devotional"
            }
    
    def _phase3_generate_prompts(self, structure: Dict, analysis: Dict, song_data: Dict, deity: str) -> Dict:
        """
        Phase 3: Generate detailed, culturally accurate prompts for each segment
        Optimized for 4K quality and Sora AI video generation
        """
        print("  ✨ Phase 3: Generating AI prompts...")
        
        segments_needing_ai = [s for s in structure['segments'] 
                               if s['visual_approach'] in ['ai_scene', 'hybrid']]
        
        if not segments_needing_ai:
            print("    ℹ️ No AI segments needed - simple video approach")
            return {
                "approach": "simple",
                "prompts": [],
                "instructions": "Use static background with audio visualizer or slideshow"
            }
        
        prompts = []
        
        for i, segment in enumerate(segments_needing_ai):
            prompt = f"""
You are an expert AI video prompt engineer specializing in culturally accurate devotional content.

Create a detailed Sora AI video prompt for this segment:

DEITY: {deity}
SEGMENT: {segment['segment_type']}
DURATION: {segment['duration_sec']} seconds
DESCRIPTION: {segment['description']}
EMOTIONAL TONE: {analysis['emotional_tone']}
CULTURAL ELEMENTS: {', '.join(segment.get('cultural_elements', []))}

PROMPT REQUIREMENTS:
1. Culturally accurate representation of {deity}
2. 4K quality descriptors (ultra HD, cinematic, high detail)
3. Appropriate for devotional/religious content
4. Specific camera movements and lighting
5. Traditional iconography and symbolism
6. Respects religious sentiments
7. Length: {segment['duration_sec']} seconds max (Sora limit)

CRITICAL PROMPT OPTIMIZATION:
- DO NOT use narrative phrases like "Opens with", "The climax of", "Begins with", "Shows"
- Start DIRECTLY with visual descriptions ("Cosmic swirling nebula", not "Opens with cosmic nebula")
- DO NOT name specific forms/avatars (Shailaputri, Brahmacharini) - instead DESCRIBE them visually
  Example: Instead of "Shailaputri form", say "Goddess in white robes holding a lotus, seated on a bull"
- Maximize descriptive tokens - every word should describe what camera sees
- Be specific about colors, textures, movements, lighting

Return JSON:
{{
  "segment_id": {i+1},
  "timestamp": "{segment['timestamp']}",
  "sora_prompt": "detailed 4K-ready prompt for Sora (pure visual descriptions, no narrative)",
  "alternative_prompt": "backup version (also pure visual)",
  "style_tags": ["cinematic", "4k", "devotional", etc],
  "camera_notes": "camera movement description",
  "lighting_notes": "lighting setup",
  "cultural_accuracy_check": "verification notes"
}}

Be descriptive but concise. Focus on visual quality and cultural authenticity. Remove all narrative framing.
"""
            
            time.sleep(self.api_call_delay)
            response = self.client.models.generate_content(model=self.model_name, contents=prompt)
            
            try:
                json_match = re.search(r'\{.*\}', response.text, re.DOTALL)
                if json_match:
                    prompt_data = json.loads(json_match.group(0))
                    prompts.append(prompt_data)
                    print(f"    ✅ Prompt {i+1}/{len(segments_needing_ai)}: {segment['timestamp']}")
            except Exception as e:
                print(f"    ⚠️ Prompt {i+1} parse error: {e}")
        
        return {
            "approach": "ai_generated",
            "prompts": prompts,
            "total_segments": len(prompts)
        }
    
    def _phase4_generate_simple_instructions(self, structure: Dict, analysis: Dict, deity: str, song_data: Dict) -> Dict:
        """
        Phase 4: For simple videos (no AI), generate implementation instructions
        Covers: slideshow, audio visualizer, static backgrounds
        """
        print("  📝 Phase 4: Creating implementation guide...")
        
        strategy = analysis['recommended_strategy']
        
        if strategy == 'simple_animated':
            instructions = {
                "type": "audio_visualizer",
                "steps": [
                    f"1. Create background: {deity} deity image or temple scene",
                    "2. Add audio waveform visualizer (bars or circular)",
                    "3. Use color palette: " + ', '.join(structure.get('color_palette', ['#FF6B35'])),
                    "4. Add subtle particle effects (optional)",
                    "5. Export in 4K (3840x2160)",
                    f"6. Duration: {structure['total_duration']}s"
                ],
                "tools": ["Adobe After Effects", "DaVinci Resolve", "Blender"],
                "assets_needed": [
                    f"{deity} deity image (high resolution)",
                    "Background texture/gradient",
                    "Audio file (.wav)"
                ]
            }
        
        elif strategy == 'slideshow':
            instructions = {
                "type": "image_slideshow",
                "steps": [
                    "1. Collect 10-15 devotional images of " + deity,
                    "2. Arrange chronologically based on song structure",
                    "3. Add smooth transitions (ken burns effect, fades)",
                    "4. Sync transitions with music tempo",
                    "5. Add text overlays for key lyrics (optional)",
                    "6. Color grade for consistent look",
                    "7. Export in 4K"
                ],
                "tools": ["Adobe Premiere Pro", "Final Cut Pro", "DaVinci Resolve"],
                "assets_needed": [
                    "10-15 HD images of " + deity,
                    "Background music",
                    "Font for text overlays (Devanagari support)"
                ]
            }
        
        else:  # hybrid or full_ai
            instructions = {
                "type": "hybrid",
                "steps": [
                    "1. Use AI-generated scenes from Phase 3 prompts",
                    "2. Fill gaps with static images/slideshows",
                    "3. Ensure smooth transitions between AI and static",
                    "4. Maintain consistent color grading",
                    "5. Add motion graphics for titles/credits",
                    "6. Export in 4K"
                ],
                "tools": ["Adobe Premiere Pro", "After Effects", "Sora AI"],
                "assets_needed": [
                    "AI-generated video clips (from Sora)",
                    "Supplementary images",
                    "Background music"
                ]
            }
        
        print(f"    ✅ Guide: {instructions['type']} ({len(instructions['steps'])} steps)")
        return instructions
    
    def _phase5_compile_final_output(self, song_data: Dict, analysis: Dict, structure: Dict, 
                                     prompts: Dict, instructions: Dict, deity: str, 
                                     day: int, user_notes: str, video_type: str) -> str:
        """
        Phase 5: Compile everything into a comprehensive text file
        """
        print("  💾 Phase 5: Compiling final output...")
        
        # Create filename
        safe_title = re.sub(r'[^\w\s-]', '', song_data['title'][:50]).strip()
        safe_title = re.sub(r'[\s]+', '_', safe_title).lower()
        filename = f"day_{day}_{deity.lower()}_{safe_title}_video_guide.txt"
        filepath = self.video_output_dir / filename
        
        # Build content
        content = []
        content.append("=" * 80)
        content.append(f"VIDEO GENERATION GUIDE - DAY {day}")
        content.append("=" * 80)
        content.append("")
        content.append(f"SONG TITLE: {song_data['title']}")
        content.append(f"DEITY: {deity}")
        content.append(f"TYPE: {video_type}")
        content.append(f"USER NOTES: {user_notes or 'None'}")
        content.append("")
        content.append("=" * 80)
        content.append("CONTENT ANALYSIS")
        content.append("=" * 80)
        content.append(f"Content Type: {analysis['content_type']}")
        content.append(f"Narrative Complexity: {analysis['narrative_complexity']}/10")
        content.append(f"Visual Density: {analysis['visual_density']}/10")
        content.append(f"AI Requirement: {analysis['ai_requirement_score']}/10")
        content.append(f"Recommended Strategy: {analysis['recommended_strategy']}")
        content.append(f"Emotional Tone: {analysis['emotional_tone']}")
        content.append(f"Reasoning: {analysis['reasoning']}")
        content.append("")
        
        if analysis.get('cultural_elements'):
            content.append(f"Cultural Elements: {', '.join(analysis['cultural_elements'])}")
        if analysis.get('key_scenes'):
            content.append(f"Key Scenes: {', '.join(analysis['key_scenes'])}")
        content.append("")
        
        content.append("=" * 80)
        content.append("VIDEO STRUCTURE")
        content.append("=" * 80)
        content.append(f"Total Duration: {structure['total_duration']} seconds")
        content.append(f"Overall Style: {structure.get('overall_style', 'Devotional')}")
        content.append(f"Color Palette: {', '.join(structure.get('color_palette', []))}")
        content.append("")
        content.append("TIMELINE:")
        content.append("-" * 80)
        
        for i, seg in enumerate(structure['segments'], 1):
            content.append(f"\n[Segment {i}] {seg['timestamp']} ({seg['duration_sec']}s)")
            content.append(f"  Type: {seg['segment_type']}")
            content.append(f"  Visual: {seg['visual_approach']}")
            content.append(f"  Description: {seg['description']}")
            if seg.get('cultural_elements'):
                content.append(f"  Cultural: {', '.join(seg['cultural_elements'])}")
            content.append(f"  Transition: {seg['transition']}")
        
        content.append("")
        content.append("")
        
        # Add prompts if AI-generated
        if prompts['approach'] == 'ai_generated' and prompts['prompts']:
            content.append("=" * 80)
            content.append("AI VIDEO PROMPTS (SORA)")
            content.append("=" * 80)
            content.append(f"Total AI Segments: {prompts['total_segments']}")
            content.append("")
            
            for p in prompts['prompts']:
                content.append(f"\n[Segment {p['segment_id']}] {p['timestamp']}")
                content.append("-" * 80)
                content.append(f"SORA PROMPT:")
                content.append(p['sora_prompt'])
                content.append("")
                content.append(f"ALTERNATIVE:")
                content.append(p.get('alternative_prompt', 'N/A'))
                content.append("")
                content.append(f"Style Tags: {', '.join(p.get('style_tags', []))}")
                content.append(f"Camera: {p.get('camera_notes', 'N/A')}")
                content.append(f"Lighting: {p.get('lighting_notes', 'N/A')}")
                content.append(f"Cultural Check: {p.get('cultural_accuracy_check', 'N/A')}")
                content.append("")
        
        # Add implementation instructions
        content.append("=" * 80)
        content.append("IMPLEMENTATION GUIDE")
        content.append("=" * 80)
        content.append(f"Approach: {instructions['type']}")
        content.append("")
        content.append("STEPS:")
        for step in instructions['steps']:
            content.append(f"  {step}")
        content.append("")
        content.append(f"Recommended Tools: {', '.join(instructions['tools'])}")
        content.append("")
        content.append("Assets Needed:")
        for asset in instructions['assets_needed']:
            content.append(f"  - {asset}")
        content.append("")
        
        content.append("=" * 80)
        content.append("SONG REFERENCE")
        content.append("=" * 80)
        content.append(f"YouTube Title: {song_data.get('youtube_title', 'N/A')}")
        content.append("")
        content.append("LYRICS EXCERPT (for timing reference):")
        content.append("-" * 80)
        content.append(song_data['lyrics'][:1000])
        if len(song_data['lyrics']) > 1000:
            content.append("\n... (truncated, see full lyrics in song file) ...")
        content.append("")
        
        content.append("=" * 80)
        content.append(f"Generated: {time.strftime('%Y-%m-%d')}")
        content.append("=" * 80)
        
        # Write file
        final_content = '\n'.join(content)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(final_content)
        
        print(f"    ✅ Saved: {filename}")
        return str(filepath)
    
    def generate_video_guide(self, row_num: int) -> Optional[str]:
        """
        Generate complete video guide for one row
        Returns: filepath of generated guide
        """
        day = self._get_cell_value(row_num, 'Day')
        video_type = self._get_cell_value(row_num, 'Type')
        deity = self._get_cell_value(row_num, 'God')
        script_file = self._get_cell_value(row_num, 'Script')
        user_notes = self._get_cell_value(row_num, 'User Notes')
        to_generate = self._get_cell_value(row_num, 'To generate')
        
        # Validation
        if not to_generate or str(to_generate).lower() not in ['true', '1', 'yes']:
            print(f"⏭️  Day {day}: Skipping (To generate = {to_generate})")
            return None
        
        if not script_file:
            print(f"⚠️  Day {day}: No script file specified")
            return None
        
        print(f"\n{'='*80}")
        print(f"🎥 DAY {day}: {deity} - {video_type}")
        print(f"{'='*80}")
        
        try:
            # Extract song content
            print(f"  📂 Loading: {script_file}")
            song_data = self._extract_lyrics_from_song_file(script_file)
            
            # Phase 1: Content Analysis
            analysis = self._phase1_analyze_content(song_data, user_notes)
            time.sleep(self.phase_delay)
            
            # Phase 2: Video Structure
            structure = self._phase2_generate_video_structure(analysis, song_data, video_type)
            time.sleep(self.phase_delay)
            
            # Phase 3: AI Prompts (if needed)
            prompts = self._phase3_generate_prompts(structure, analysis, song_data, deity)
            time.sleep(self.phase_delay)
            
            # Phase 4: Implementation Instructions
            instructions = self._phase4_generate_simple_instructions(structure, analysis, deity, song_data)
            time.sleep(self.phase_delay)
            
            # Phase 5: Compile Final Output
            filepath = self._phase5_compile_final_output(
                song_data, analysis, structure, prompts, instructions,
                deity, day, user_notes or '', video_type
            )
            
            # Update Excel
            self._set_cell_value(row_num, 'Video', filepath)
            
            print(f"\n✅ Day {day} Complete!")
            print(f"📁 Output: {filepath}")
            
            return filepath
            
        except FileNotFoundError as e:
            print(f"❌ Error: {e}")
            return None
        except Exception as e:
            print(f"❌ Unexpected error: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_all(self, start_day: Optional[int] = None, end_day: Optional[int] = None):
        """
        Generate video guides for all marked rows
        """
        print(f"\n{'='*80}")
        print("🎬 VIDEO GENERATION PIPELINE - STARTING")
        print(f"{'='*80}\n")
        
        generated = []
        skipped = []
        errors = []
        
        for row_num in range(2, self.ws.max_row + 1):
            day = self._get_cell_value(row_num, 'Day')
            
            if day is None:
                continue
            
            # Filter by day range if specified
            if start_day and day < start_day:
                continue
            if end_day and day > end_day:
                continue
            
            result = self.generate_video_guide(row_num)
            
            if result:
                generated.append((day, result))
            elif result is None:
                skipped.append(day)
            else:
                errors.append(day)
            
            # Delay between videos
            if result and row_num < self.ws.max_row:
                print(f"\n⏳ Waiting {self.video_delay}s before next video...")
                time.sleep(self.video_delay)
        
        # Summary
        print(f"\n{'='*80}")
        print("📊 GENERATION SUMMARY")
        print(f"{'='*80}")
        print(f"✅ Generated: {len(generated)}")
        print(f"⏭️  Skipped: {len(skipped)}")
        print(f"❌ Errors: {len(errors)}")
        
        if generated:
            print(f"\nGenerated Days: {', '.join([str(d) for d, _ in generated])}")
        
        print(f"\n{'='*80}")
        print("🎉 VIDEO GENERATION COMPLETE!")
        print(f"{'='*80}\n")


def main():
    """Main entry point"""
    import sys
    
    excel_path = "First30DaysVideo.xlsx"
    
    # Parse command line arguments
    start_day = None
    end_day = None
    
    if len(sys.argv) >= 2:
        start_day = int(sys.argv[1])
        end_day = int(sys.argv[2]) if len(sys.argv) >= 3 else start_day
    
    try:
        generator = VideoPromptGenerator(excel_path)
        generator.generate_all(start_day, end_day)
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
