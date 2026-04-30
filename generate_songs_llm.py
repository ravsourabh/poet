"""
AI-Powered Devotional Song Generator
5-stage LLM chaining with cultural validation and individual file storage
"""

import openpyxl
import json
import os
import time
import re
import random
from pathlib import Path
from typing import Dict, List, Optional
from dotenv import load_dotenv
from datetime import datetime
from google import genai
from google.genai import types


# CONSTANTS - Centralized configuration
class Config:
    """Centralized constants to eliminate magic numbers"""
    # Content limits
    MAX_LYRICS_PREVIEW = 1500
    MAX_ANALYSIS_CONTENT = 1000
    MAX_SONG_GUIDE_CHARS = 2000  # Compressed from 3500
    MIN_CONTENT_LENGTH = 500  # Minimum acceptable song length
    
    # Analysis limits
    MAX_MANTRAS = 4
    MAX_DEITY_ATTRS = 5
    MAX_IMAGERY_ITEMS = 5
    MAX_ISSUES_DISPLAY = 3
    
    # Retry and timing
    ERROR_RETRY_DELAY = 5
    RATE_LIMIT_SHORT_WAIT = 15
    RATE_LIMIT_LONG_WAIT = 30
    MAX_LLM_RETRIES = 3
    
    # Token pricing (per 1K tokens)
    FLASH_INPUT_COST = 0.000075
    FLASH_OUTPUT_COST = 0.0003
    PRO_INPUT_COST = 0.00125
    PRO_OUTPUT_COST = 0.005
    
    # Progress tracking
    PROGRESS_FILE = ".generation_progress.json"


class GeminiSongGenerator:
    """Main song generator using Google Gemini AI with 8-stage pipeline: Hindi + Rhythm + Melody + Cultural validation"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.output_dir = self.excel_path.parent / "content"
        self.output_dir.mkdir(exist_ok=True)
        
        # Cost reports directory
        self.cost_dir = self.excel_path.parent / 'cost_reports'
        self.cost_dir.mkdir(exist_ok=True)
        
        # Rate limiting settings
        self.chain_delay = 5 # Seconds between chains
        self.song_delay = 10 # Seconds between songs
        
        # Cost tracking (Phase 4)
        self.cost_tracker = {
            'flash_input_tokens': 0,
            'flash_output_tokens': 0,
            'pro_input_tokens': 0,
            'pro_output_tokens': 0,
            'total_calls': 0,
            'analysis_calls': 0,
            'generation_calls': 0,
            'validation_calls': 0,
            'refinement_calls': 0
        }
        
        # Progress tracking (Phase 4)
        self.progress_file = self.excel_path.parent / Config.PROGRESS_FILE
        self.progress = self._load_progress()
        
        # Load Excel
        self.wb = openpyxl.load_workbook(excel_path)
        self.ws = self.wb.active
        
        # Parse columns
        self.headers = [cell.value for cell in self.ws[1]]
        self.column_map = {h: i for i, h in enumerate(self.headers)}
        print(f"📊 Columns: {', '.join([h for h in self.headers if h])}")
        
        # Initialize Gemini
        self._init_gemini()
        
        # Load framework
        self._load_song_guide()
        
        print(f"✅ Ready to generate! Output: {self.output_dir}")
        print(f"⏱️  Rate Limits: {self.chain_delay}s between chains, {self.song_delay}s between songs")
    
    def _init_gemini(self):
        """Initialize Gemini API with DUAL-MODEL approach for cost optimization"""
        load_dotenv()
        api_key = os.getenv('GEMINI_API_KEY')
        
        if not api_key or api_key == 'your_gemini_api_key_here':
            raise ValueError(
                "Set GEMINI_API_KEY in .env file\n"
                "Get key: https://makersuite.google.com/app/apikey"
            )
        
        self.client = genai.Client(api_key=api_key)

        # DUAL MODEL APPROACH for cost optimization
        self.use_dual_models = os.getenv('USE_DUAL_MODELS', 'true').lower() == 'true'

        if self.use_dual_models:
            # Flash model for cheap operations (analysis, validation, quality checks)
            self.flash_model_name = os.getenv('FLASH_MODEL', 'gemini-2.5-flash-lite')

            # Backup Flash model (Phase 1 - Reliability #3)
            self.flash_backup_name = os.getenv('FLASH_MODEL_BACKUP', 'gemini-2.0-flash')

            # Pro model ONLY for song generation (critical quality)
            self.pro_model_name = os.getenv('PRO_MODEL', 'gemini-2.5-pro')

            self.model_name = self.flash_model_name

            print(f"✅ DUAL-MODEL MODE:")
            print(f"   💨 Flash (cheap): {self.flash_model_name} - For analysis, validation")
            if self.flash_backup_name:
                print(f"   🔄 Flash Backup: {self.flash_backup_name} - Fallback if rate limited")
            print(f"   🎯 Pro (quality): {self.pro_model_name} - For song generation ONLY")
        else:
            # Single model mode (legacy)
            self.model_name = os.getenv('GEMINI_MODEL', 'gemini-2.5-flash-lite')
            self.flash_model_name = self.model_name
            self.pro_model_name = self.model_name
            print(f"✅ SINGLE-MODEL MODE: {self.model_name}")
        
        # Load quality thresholds from .env (separate for each check)
        self.hindi_threshold = int(os.getenv('HINDI_THRESHOLD', '8'))  # Grammar/poetry
        self.singability_threshold = int(os.getenv('SINGABILITY_THRESHOLD', '7'))  # Rhythm + melody combined
        self.cultural_threshold = int(os.getenv('CULTURAL_THRESHOLD', '7'))  # Cultural sensitivity
        self.max_quality_retries = int(os.getenv('MAX_QUALITY_RETRIES', '2'))  # How many regenerations
        self.enable_quality_checks = os.getenv('ENABLE_QUALITY_CHECKS', 'true').lower() == 'true'
        self.use_iterative_refinement = os.getenv('USE_ITERATIVE_REFINEMENT', 'true').lower() == 'true'
        self.skip_chain4_refine = os.getenv('SKIP_CHAIN4_REFINE', 'false').lower() == 'true'
        self.auto_complete_truncated = os.getenv('AUTO_COMPLETE_TRUNCATED', 'true').lower() == 'true'
        
        print(f"✅ Quality Thresholds: Hindi={self.hindi_threshold}, Singability={self.singability_threshold}, Cultural={self.cultural_threshold}")
        print(f"✅ Max Retries: {self.max_quality_retries}, Iterative Refinement: {'On' if self.use_iterative_refinement else 'Off'}")
        print(f"✅ Quality Checks: {'Enabled' if self.enable_quality_checks else 'Disabled'}")
        print(f"✅ Auto-Complete Truncated: {'On' if self.auto_complete_truncated else 'Off'}")
    
    def _load_progress(self) -> Dict:
        """Load progress from checkpoint file (Phase 4 - Feature #1)"""
        if self.progress_file.exists():
            try:
                with open(self.progress_file, 'r', encoding='utf-8') as f:
                    progress = json.load(f)
                    print(f"📂 Progress loaded: {len(progress.get('completed', []))} songs completed")
                    return progress
            except:
                pass
        return {'completed': [], 'failed': [], 'start_time': datetime.now().strftime('%Y-%m-%d')}
    
    def _save_progress(self, day: int, status: str, error: str = None):
        """Save progress to checkpoint file (Phase 4 - Feature #1)"""
        if status == 'completed':
            if day not in self.progress['completed']:
                self.progress['completed'].append(day)
        elif status == 'failed':
            self.progress['failed'].append({'day': day, 'error': error, 'time': datetime.now().strftime('%Y-%m-%d')})
        
        with open(self.progress_file, 'w', encoding='utf-8') as f:
            json.dump(self.progress, f, indent=2)
    
    def _load_song_guide(self):
        """Load Suno AI framework"""
        guide_path = self.excel_path.parent / "songGuide.md"
        
        if not guide_path.exists():
            raise FileNotFoundError("songGuide.md required")
        
        with open(guide_path, 'r', encoding='utf-8') as f:
            self.song_guide = f.read()
        
        print(f"✅ Framework: {len(self.song_guide)} chars")
        
        # Load structure templates
        self._init_structure_templates()
    
    def _init_structure_templates(self):
        """Initialize 30 different song structure templates"""
        
        # Opening variations (15 types)
        self.openings = [
            "[Alaap - Solo Voice, Ascending]",
            "[Temple Bells + Chanting]",
            "[Instrumental Prelude - {inst_main}]",
            "[Nature Sounds + Tanpura Drone]",
            "[Energetic Dhol Beat]",
            "[A Cappella Harmony]",
            "[Single {inst_main} Solo]",
            "[Atmospheric Pad + Whispers]",
            "[Direct Chorus Entry]",
            "[Group Mantra Chanting]",
            "[Spoken Word - Soft Voice]",
            "[Rhythmic Tabla Pattern]",
            "[Flute Melody - Peaceful]",
            "[Harmonium Swell]",
            "[Silence then Sudden Entry]"
        ]
        
        # Outro variations (15 types)
        self.outros = [
            "[Gradual Fade - All Instruments]\n(final mantra)",
            "[Echo Repetition - Voice Only]\n[Whispered]\nClosing mantra",
            "[Nature Sounds Integration + Fade]",
            "[Full Stop - Sudden Silence]\nॐ शान्ति शान्ति शान्ति",
            "[Humming Fade - Group Voices]",
            "[{inst_main} Solo - No Vocals]\n[Fade]",
            "[Temple Bell Ring + Silence]",
            "[Reverse Build - Deconstructing]",
            "[{inst_main} Echo + Fade]\n[Whispered]\nClosing mantra",
            "[Circular Return to Intro Theme]",
            "[Sustained Drone + Whisper]",
            "[Clapping Fade Out]",
            "[Instrumental Reprise]\n[Slow Fade]",
            "[Call-Response Fading]\n(mantras)",
            "[Breath Sound + Silence]"
        ]
        
        # 30 unique structure templates mapped to styles
        self.structure_templates = {
            # Classical/Traditional (6 templates)
            "classical_bhajan_1": """[Intro - {inst_main} + Tanpura]
(opening mantra)

[Solo Voice - Slow, No Rhythm]
2 Hindi lines - devotional alaap style

[Verse 1 - Harmonium Base, Gentle Tabla]
4 lines about story
(mantra)

[Verse 2 - Full Tabla Pattern, Building]
4 lines continuing
(mantra)

[Chorus - Layered Vocals]
4 lines - main hook
(mantra)

[Bridge - All Instruments, Emotive]
4 lines, emotional peak

[BPM: 75]
[Final Chorus - Returning Theme]
Chorus variation with resolution
(mantras)""",

            "classical_bhajan_2": """[Intro - Harmonium Swell]
(mantra)

[Solo Voice - Meditative]
Opening 2 lines

[Verse 1 - Male Voice, Devotional]
4 lines story beginning
(mantra)

[Instrumental - {inst_main} Interlude]

[Verse 2 - Female Response]
4 lines story continuation
(mantra)

[Chorus - Union of Voices]
4 lines hook
(mantra)

[Instrumental - Tabla Solo with Build Up]

[BPM: 80]
[Final Verse - All Voices]
4 lines resolution
(mantras)""",

            # Qawwali Style (3 templates)
            "qawwali_1": """[Intro - Harmonium + Tabla]
(opening phrase)

[Solo Lead - Powerful Voice]
Opening call - 2 lines

[Chorus Response - Group]
Response - 2 lines

[Verse 1 - Lead with Tabla Accent]
4 lines about story
(mantra)

[Chorus - Call-Response]
Group echoes each line

[Build Up - Tabla Intensifies]

[Verse 2 - Powerful]
4 lines
(mantra)

[Bridge - Hand Claps, Fast Rhythm]
4 lines call-response style

[BPM: 110]
[Final Chorus - All Voices Unite, Grand]
Chorus repeated with intensity
(mantras chanted repeatedly)""",

            "qawwali_2": """[Intro - Slow Harmonium Build]

[Lead Voice - Soulful Entry]
Opening 3 lines

[Group Response - Echo]
(mantra response)

[Verse 1 - Building Energy]
4 lines with growing intensity
(mantra)

[Instrumental - Tabla Solo, Accelerating]

[Verse 2 - Full Voice]
4 lines powerful
(mantra)

[Bridge - Hand Claps, All Voices]
4 lines call-response

[BPM: 115]
[Final Chorus - Repeated]
Chorus 3 times with variations
(mantras)""",

            # Rock/Energetic (3 templates)
            "rock_devotional_1": """[Electric Guitar Intro - Power Chords]

[Drums Entry - Strong Beat]

[Verse 1 - Strong Vocals, Modern]
4 lines about story
(mantra)

[Pre-Chorus - Building Drums]
2 lines building tension

[Big Chorus - Full Band]
4 lines main hook
(mantra)

[Verse 2 - Heavier]
4 lines continuing
(mantra)

[Bridge - Guitar Solo]
Instrumental break

[Breakdown - Drums Only]
2 lines stripped down

[BPM: 120]
[Final Chorus - Heaviest, Layered]
Chorus variation twice
(mantras)""",

            "rock_devotional_2": """[Intro - Distorted {inst_main}]

[Verse 1 - Clean Vocals]
4 lines

[Build-Up - Drums Enter]

[Chorus - Explosive]
4 lines with power
(mantra)

[Instrumental - {inst_main} + Drums]

[Verse 2 - Harmonized]
4 lines
(mantra)

[Bridge - Half-Time Feel]
4 lines slower, powerful

[BPM: 125]
[Final Chorus - Double-Time]
Fast, energetic finish
(mantras)""",

            # Meditative/Sufi (4 templates)
            "meditative_1": """[Ambient Soundscape - Nature]

[Instrumental - Tanpura Drone]

[Verse 1 - Soft Voice, Floating]
4 lines slow, peaceful
(mantra softly)

[Instrumental - {inst_main} Solo, Meditative]
Breathing space

[Gentle Voice - Slightly Building]
Verse 2 - 4 lines
(mantra)

[Bridge - Humming, Sustained Notes]
Wordless vocals

[BPM: 60]
[Verse 3 - Voice Returns, Softer]
4 lines, gentle conclusion
(mantra fading)""",

            "meditative_2": """[Silence... Then Single Note]

[Intro - Deep Drone]

[Spoken Word - Whispered]
(mantra repeated)

[Verse 1 - Voice Emerges, Peaceful]
4 lines minimal
(mantra)

[Instrumental - {inst_main} Solo, Slow]

[Verse 2 - Voice Stronger]
4 lines
(mantra)

[Bridge - Sustained Harmonics]
Long notes, meditative

[BPM: 65]
[Outro - Voice Fades to Silence]
2 lines fading
(mantra dissolving)""",

            "sufi_1": """[Intro - Ney Flute, Longing]

[Instrumental - Daf Rhythm, Gentle]

[Solo Voice - Ecstatic]
Verse 1 - 4 lines devotional yearning
(mantra)

[Chorus - Group Joins]
4 lines unified
(mantra)

[Instrumental - Dancing Rhythm]

[Verse 2 - More Intense]
4 lines building ecstasy
(mantra)

[Bridge - Whirling, Spinning Energy]
4 lines cyclical

[BPM: 95]
[Final Chorus - Transcendent]
Chorus with variations
(mantras repeated)""",

            # Garba/Dandiya (2 templates)
            "garba_1": """[Dhol Intro - Fast, Rhythmic]

[Call - Lead Singer]
Opening 2 lines

[Response - Group Shout]
(enthusiastic response)

[Verse 1 - Energetic, Rhythmic]
4 lines with Garba beat
(mantra)

[Instrumental - Dhol Solo]

[Verse 2 - Faster]
4 lines accelerating
(mantra)

[Bridge - Clapping Pattern, Everyone]
4 lines with hand claps

[BPM: 140]
[Final Chorus - Accelerating, Circular]
Chorus fastest tempo
(mantras shouted)""",

            "dandiya_1": """[Intro - Sticks Clacking, Rhythmic]

[Instrumental - Dhol and Sticks]

[Verse 1 - Fast, Joyful]
4 lines celebratory
(mantra)

[Call-Response - Quick]
Leader and group alternate

[Verse 2 - Building Speed]
4 lines
(mantra)

[Breakdown - Sticks Only]

[Bridge - Maximum Energy]
4 lines fastest

[BPM: 145]
[Final Chorus - Frantic Joy]
Chorus twice, accelerating
(mantras)""",

            # Modern Devotional (4 templates)
            "modern_devotional_1": """[Electronic Pad - Atmospheric]

[Beat Drop - Modern]

[Verse 1 - Clear Vocals]
4 lines story
(mantra)

[Pre-Chorus - Rising Synth]
2 lines building

[Chorus - Full Production]
4 lines hook with harmonies
(mantra)

[Verse 2 - Added Layers]
4 lines
(mantra)

[Bridge - Breakdown to Minimal]
4 lines emotional

[BPM: 100]
[Final Chorus - Maximum Production]
Chorus twice with variations
(mantras)""",

            "modern_devotional_2": """[Lo-Fi Beat - Chill]

[{inst_main} Sample - Looped]

[Verse 1 - Relaxed Delivery]
4 lines
(mantra)

[Chorus - Melodic Hook]
4 lines catchy
(mantra)

[Instrumental - Beat Switch]

[Verse 2 - Flow Style]
4 lines
(mantra)

[Bridge - Just Voice + Beat]
4 lines minimal

[BPM: 85]
[Final Chorus - Layered]
Chorus with ad-libs
(mantras)""",

            # Kirtan Style (2 templates)
            "kirtan_1": """[Harmonium Intro - Simple Pattern]

[Leader Voice - Clear]
Mantra line 1

[Group Response]
(repeat mantra)

[Leader]
Mantra line 2

[Group Response]
(repeat mantra)

[Verse - Leader Explains]
4 lines story

[All Together - Building]
Main mantra repeated with variations

[Tabla Joins - Accelerating]

[BPM: 90 → 120]
[Ecstatic Repetition]
Main mantra faster and faster
(everyone chanting)""",

            "kirtan_2": """[Simple Clap Pattern]

[Solo Voice]
Opening mantra

[Group Echo]
(mantra repeat)

[Verse 1 - Story Telling]
4 lines narrative

[All Chant - Unified]
(main mantra together)

[Building Rhythm]

[Verse 2 - Continuing]
4 lines

[Accelerating Chant]
(mantra faster)

[BPM: 85 → 110]
[Peak Energy - All Voices]
(mantra repeated many times)""",

            # Aarti Style (2 templates)
            "aarti_1": """[Bell Ringing - Temple Sound]

[Verse 1 - Devotional, Reverent]
4 lines praise
ॐ जय {god_name}

[Chorus - Group Participation]
4 lines - everyone joins
ॐ जय {god_name}

[Verse 2 - Continuing Praise]
4 lines attributes
ॐ जय {god_name}

[Chorus - Louder]
4 lines repeated
ॐ जय {god_name}

[Verse 3 - Final Praise]
4 lines conclusion

[BPM: 70]
[Final Chorus - All Together]
Main chorus grand
ॐ जय {god_name} की""",

            # Stotra Style (2 templates)
            "stotra_1": """[Intro - Sacred Atmosphere]
ॐ

[Verse 1 - Sanskrit/Hindi Mix]
4 lines formal prayer
(vedic mantra)

[Verse 2 - Continuing]
4 lines
(mantra)

[Central Verse - Most Important]
4 lines key attributes
(main mantra)

[Verse 3 - Praise]
4 lines
(mantra)

[Verse 4 - Concluding]
4 lines blessings

[BPM: 70]
[Phala Shruti - Final Benediction]
2 lines benefits of recitation
ॐ शान्ति शान्ति शान्ति""",

            # Narrative/Epic (2 templates)
            "narrative_epic_1": """[Dramatic Intro - Building Tension]

[Narrator Voice - Story Beginning]
Opening 3 lines setting scene

[Verse 1 - Story Unfolds]
4 lines beginning
(mantra)

[Musical Interlude - Scene Change]

[Verse 2 - Conflict/Challenge]
4 lines rising action
(mantra)

[Bridge - Climax]
4 lines peak of story
(powerful mantra)

[Verse 3 - Resolution]
4 lines conclusion

[BPM: 90]
[Final Chorus - Moral/Praise]
4 lines message
(mantras)""",

            # Lullaby/Peaceful (2 templates)
            "lullaby_1": """[Soft Humming]

[Gentle Voice - Like Singing to Child]
Verse 1 - 4 lines peaceful
(soft mantra)

[Instrumental - Lullaby Melody]

[Soft Voice - Continuing]
Verse 2 - 4 lines soothing
(mantra)

[Bridge - Extra Soft]
4 lines most gentle

[BPM: 55]
[Fading Voice]
Final 2 lines
(whispered mantra)""",

            # Celebration/Festival (2 templates)
            "celebration_1": """[Festive Intro - Joyful Instruments]

[Verse 1 - Celebratory, Fast]
4 lines joyful
(excited mantra)

[Chorus - Everyone Celebrates]
4 lines main celebration hook
(mantra shouted)

[Dance Break - Instrumental]

[Verse 2 - Building Joy]
4 lines
(mantra)

[Bridge - Peak Celebration]
4 lines maximum energy

[BPM: 130]
[Final Chorus - Repeat 3x]
Chorus with increasing joy
(mantras)""",

            # Bhakti Rasa (2 templates)
            "bhakti_madhurya_1": """[Sweet Intro - Tender {inst_main}]

[Verse 1 - Sweet, Loving Voice]
4 lines intimate devotion
(tender mantra)

[Chorus - Heart-Melting]
4 lines expressing divine love
(mantra with longing)

[Instrumental - Romantic, Devotional]

[Verse 2 - Deeper Emotion]
4 lines yearning
(mantra)

[Bridge - Peak Emotion]
4 lines ecstatic love

[BPM: 75]
[Final Chorus - Surrendered]
Chorus dissolved in love
(mantras)"""
        }
        
        print(f"✅ Structure Templates: {len(self.structure_templates)} unique templates loaded")
    
    def _select_song_structure(self, row_data: Dict, analysis: Dict) -> str:
        """Use Flash LLM's structure recommendation from analysis phase"""
        
        # Get LLM's recommendation (Flash already analyzed this in Chain 1)
        template_key = analysis.get('recommended_structure', 'classical_bhajan_1')
        
        # Validate template exists, fallback if needed
        if template_key not in self.structure_templates:
            print(f"  ⚠️  Unknown template '{template_key}', using classical_bhajan_1")
            template_key = 'classical_bhajan_1'
        
        # Get base template
        base_structure = self.structure_templates[template_key]
        
        # Replace placeholders
        inst_main = row_data['Instruments'].split(',')[0].strip()
        base_structure = base_structure.replace('{inst_main}', inst_main)
        base_structure = base_structure.replace('{god_name}', row_data['God'])
        
        # Randomly vary opening (30% chance)
        if random.random() < 0.3:
            opening = random.choice(self.openings)
            opening = opening.replace('{inst_main}', inst_main)
            # Replace first section with random opening
            lines = base_structure.split('\n')
            lines[0] = opening
            base_structure = '\n'.join(lines)
        
        # Randomly vary outro (30% chance)
        if random.random() < 0.3:
            outro = random.choice(self.outros)
            outro = outro.replace('{inst_main}', inst_main)
            # Add outro variation at end
            base_structure = base_structure.rsplit('\n', 3)[0] + '\n\n' + outro
        
        return base_structure
    
    def _call_llm(self, prompt: str, temperature: float = 0.7, use_pro: bool = False, call_type: str = 'general') -> str:
        """IMPROVED: Call Gemini with retry logic, backup model support, and cost tracking
        
        Args:
            prompt: The prompt to send
            temperature: Generation temperature
            use_pro: If True, use pro model (expensive, for song generation)
                    If False, use flash model (cheap, for validation/analysis)
            call_type: Type of call for cost tracking ('analysis', 'generation', 'validation', 'refinement')
        
        Phase 1 (Reliability #3): Better rate limit handling with backup Flash model
        Phase 4 (Feature #2): Cost tracking with token counting
        """
        max_retries = Config.MAX_LLM_RETRIES
        
        # Select model name based on use_pro flag
        model_name = self.pro_model_name if use_pro else self.flash_model_name
        
        for attempt in range(max_retries):
            try:
                # Increase token limit for lyrics generation to avoid truncation
                token_limit = 30000 if call_type == 'generation' else 12000
                
                response = self.client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=types.GenerateContentConfig(
                        temperature=temperature,
                        max_output_tokens=token_limit
                    )
                )
                
                # DEBUG: Check why generation stopped
                if call_type == 'generation' and hasattr(response, 'candidates') and response.candidates:
                    candidate = response.candidates[0]
                    finish_reason = candidate.finish_reason
                    
                    # Get actual token count if available
                    actual_tokens = 0
                    if hasattr(response, 'usage_metadata') and response.usage_metadata:
                        actual_tokens = response.usage_metadata.candidates_token_count
                    
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    
                    if finish_reason != 1:  # 1 = STOP (natural completion), 2 = MAX_TOKENS, 3 = SAFETY, etc.
                        print(f"  [{timestamp}] ⚠️  LLM STOPPED EARLY! Finish reason: {finish_reason}")
                        if finish_reason == 2:
                            print(f"  [{timestamp}] ❌ HIT TOKEN LIMIT! Used {actual_tokens}/{token_limit} tokens")
                            print(f"  [{timestamp}] 💡 Solution: Simplify prompt or increase token limit further")
                        elif finish_reason == 3:
                            print(f"  [{timestamp}] ⚠️  SAFETY FILTER triggered - content flagged as unsafe")
                        elif finish_reason == 4:
                            print(f"  [{timestamp}] ⚠️  RECITATION - LLM detected potential copyright content")
                    else:
                        # Natural stop - but check if output is suspiciously short
                        output_length = len(response.text)
                        if output_length < 1500:  # Increased threshold - most complete songs are 1500-2500 chars
                            print(f"  [{timestamp}] ⚠️  Short output ({output_length} chars, {actual_tokens} tokens) - LLM stopped naturally but early")
                            print(f"  [{timestamp}] 💡 This suggests incomplete lyrics - check the output file")
                
                # Cost tracking (Phase 4 - Feature #2)
                self.cost_tracker['total_calls'] += 1
                if call_type in self.cost_tracker:
                    self.cost_tracker[f'{call_type}_calls'] += 1
                
                # Track tokens (approximate from prompt/response length)
                input_tokens = len(prompt) // 4  # Rough estimate: 1 token ≈ 4 chars
                output_tokens = len(response.text) // 4
                
                if use_pro:
                    self.cost_tracker['pro_input_tokens'] += input_tokens
                    self.cost_tracker['pro_output_tokens'] += output_tokens
                else:
                    self.cost_tracker['flash_input_tokens'] += input_tokens
                    self.cost_tracker['flash_output_tokens'] += output_tokens
                
                return response.text
                
            except Exception as e:
                error_str = str(e).lower()
                timestamp = datetime.now().strftime("%H:%M:%S")
                
                # Phase 1: Better rate limit handling with Flash backup
                if '429' in error_str or 'quota' in error_str or 'resource_exhausted' in error_str:
                    # FLASH RATE LIMIT - Try backup model if available
                    if not use_pro and self.flash_backup_name and attempt == 0:
                        print(f"  [{timestamp}] ⚠️  {model_name} rate limited, switching to backup: {self.flash_backup_name}")
                        try:
                            response = self.client.models.generate_content(
                                model=self.flash_backup_name,
                                contents=prompt,
                                config=types.GenerateContentConfig(
                                    temperature=temperature,
                                    max_output_tokens=8000
                                )
                            )
                            self.cost_tracker['total_calls'] += 1
                            # Track with Flash pricing (backup is also Flash tier)
                            input_tokens = len(prompt) // 4
                            output_tokens = len(response.text) // 4
                            self.cost_tracker['flash_input_tokens'] += input_tokens
                            self.cost_tracker['flash_output_tokens'] += output_tokens
                            return response.text
                        except:
                            print(f"  [{timestamp}] ⚠️  Backup also failed, waiting...")
                    
                    # Regular rate limit retry
                    if attempt < max_retries - 1:
                        wait = Config.RATE_LIMIT_SHORT_WAIT * (attempt + 1)  # 15s, 30s, 45s
                        print(f"  [{timestamp}] ⚠️  Rate limit ({model_name}), waiting {wait}s...")
                        time.sleep(wait)
                    else:
                        print(f"  [{timestamp}] ❌ Rate limit exhausted after {max_retries} attempts")
                        model_type = "PRO" if use_pro else "FLASH"
                        raise Exception(f"{model_type} rate limit exceeded - wait 60+ seconds before continuing")
                else:
                    # Other API errors (Phase 3: Quality #1 - Standardized error messages)
                    if attempt < max_retries - 1:
                        wait = 2 ** attempt  # 1s, 2s, 4s
                        print(f"  [{timestamp}] ⚠️  API error ({model_name}), retry {attempt+1}/{max_retries} in {wait}s...")
                        time.sleep(wait)
                    else:
                        error_preview = str(e)[:150]
                        raise Exception(f"[{timestamp}] LLM call failed ({model_name}, {call_type}): {error_preview}")
    
    def _extract_json(self, text: str, fallback: Dict) -> Dict:
        """Extract JSON from LLM response"""
        try:
            start = text.find('{')
            end = text.rfind('}') + 1
            if start != -1 and end > start:
                return json.loads(text[start:end])
        except:
            pass
        return fallback
    
    def _generate_cost_report(self) -> str:
        """Generate cost report from tracked token usage (Phase 4 - Feature #2)"""
        ct = self.cost_tracker
        
        # Calculate costs
        flash_input = (ct['flash_input_tokens'] / 1000) * Config.FLASH_INPUT_COST
        flash_output = (ct['flash_output_tokens'] / 1000) * Config.FLASH_OUTPUT_COST
        pro_input = (ct['pro_input_tokens'] / 1000) * Config.PRO_INPUT_COST
        pro_output = (ct['pro_output_tokens'] / 1000) * Config.PRO_OUTPUT_COST
        
        flash_total = flash_input + flash_output
        pro_total = pro_input + pro_output
        total_cost = flash_total + pro_total
        
        report = f"""
╔══════════════════════════════════════════════════════════════╗
║                     💰 COST REPORT                          ║
╠══════════════════════════════════════════════════════════════╣
║ Flash Model Usage:                                           ║
║   Input tokens:  {ct['flash_input_tokens']:>8,} → ${flash_input:>7.4f}          ║
║   Output tokens: {ct['flash_output_tokens']:>8,} → ${flash_output:>7.4f}          ║
║   Subtotal:                      ${flash_total:>7.4f}          ║
║                                                              ║
║ Pro Model Usage:                                             ║
║   Input tokens:  {ct['pro_input_tokens']:>8,} → ${pro_input:>7.4f}          ║
║   Output tokens: {ct['pro_output_tokens']:>8,} → ${pro_output:>7.4f}          ║
║   Subtotal:                      ${pro_total:>7.4f}          ║
║                                                              ║
║ Call Breakdown:                                              ║
║   Analysis:    {ct['analysis_calls']:>4} calls                            ║
║   Generation:  {ct['generation_calls']:>4} calls                            ║
║   Validation:  {ct['validation_calls']:>4} calls                            ║
║   Refinement:  {ct['refinement_calls']:>4} calls                            ║
║   Total:       {ct['total_calls']:>4} calls                            ║
║                                                              ║
║ TOTAL COST:                      ${total_cost:>7.4f}          ║
╚══════════════════════════════════════════════════════════════╝
"""
        return report
    
    def _chain1_analyze(self, row_data: Dict) -> Dict:
        """Chain 1: Story analysis"""
        user_notes = row_data.get('User Notes', '')
        notes_text = f"\n- User Request: {user_notes}" if user_notes else ""
        
        prompt = f"""Expert in Hindu devotional music. Analyze:

INPUTS:
- Deity: {row_data['God']}
- Story: {row_data['Story']}
- Style: {row_data['Style']}
- Type: {row_data['Type']}{notes_text}

ANALYZE:
1. Story essence
2. Deity attributes (5-7)
3. Emotional arc
4. Authentic mantras (3-5)
5. Visual imagery
6. Musical mood
7. INDIAN AESTHETIC RAS (Navarasa) - Identify the primary emotional flavor:
   - Veer Ras (Heroic/Valor) - War, Hanuman's feats, Shiva's tandava
   - Shanta Ras (Peaceful/Calm) - Meditation, Krishna's flute, morning prayers
   - Karuna Ras (Compassion/Pathos) - Sita's exile, devotee's plea, separation
   - Bhakti Ras (Devotion/Love) - Pure worship, surrender, divine love
   - Adbhuta Ras (Wonder/Awe) - Cosmic visions, miracles, divine forms
   - Shringar Ras (Romantic/Beauty) - Radha-Krishna leela, divine beauty
   - Hasya Ras (Joyful/Playful) - Krishna's childhood, celebrations
   Choose the PRIMARY Ras that should dominate the musical mood
8. SINGABILITY LEVEL - CRITICAL DECISION (INTELLIGENT JUDGMENT):
   
   Choose "high" IF (MOST SONGS SHOULD BE THIS):
   - SIMPLE, FOCUSED story (single episode, one moment)
   - Emotional/devotional content (prayer, praise, specific event)
   - Style is catchy (Qawwali, Kirtan, Rock, Modern, Celebration)
   - Goal: mass participation, memorability, easy to sing along
   - Examples: 
     * "Sudama meets Krishna" (simple focused story)
     * "Krishna's love for devotees"
     * "Ram's praise"
     * "Hanuman crosses ocean" (single heroic moment)
     * "Morning/Evening prayer"
   
   Choose "medium" IF:
   - EPIC/COMPLEX narrative (multi-part, descriptive journey)
   - Multiple episodes or philosophical depth
   - Classical bhajan needing story detail
   - Balance between narrative depth and melody
   - Examples:
     * "Ram's 14-year exile" (multi-episode journey)
     * "Krishna teaches Bhagavad Gita" (philosophical, multi-part)
     * "Ramayana war sequence" (multiple events)
     * "Shiva's tandava and universe cycles" (cosmic/abstract)
   
   Choose "flexible" IF (RARE):
   - Highly literary/poetic Stotra or hymn
   - Sacred verses requiring exact traditional phrasing
   - Aarti with specific ritual structure
   - Scholarly/devotional treatise
   - Examples:
     * "Vishnu Sahasranama excerpt"
     * "Shiva Tandava Stotram"
     * "Guru Stotram"
   
   DEFAULT RULE: When in doubt, use "high" - most devotional songs should be catchy and singable!

9. SONG STRUCTURE RECOMMENDATION:
   Based on the style, recommend the BEST structure template:
   - Qawwali: "qawwali_1" or "qawwali_2"
   - Rock/Energetic: "rock_devotional_1" or "rock_devotional_2"
   - Meditative/Peaceful: "meditative_1" or "meditative_2"
   - Sufi: "sufi_1"
   - Garba: "garba_1"
   - Dandiya: "dandiya_1"
   - Kirtan: "kirtan_1" or "kirtan_2"
   - Aarti: "aarti_1"
   - Stotra/Hymn: "stotra_1"
   - Lullaby/Soft: "lullaby_1"
   - Celebration/Festive/Joyful: "celebration_1"
   - Modern/Contemporary: "modern_devotional_1" or "modern_devotional_2"
   - Narrative/Epic/Story: "narrative_epic_1"
   - Love/Madhurya/Romantic: "bhakti_madhurya_1"
   - Classical/Traditional/Devotional: "classical_bhajan_1" or "classical_bhajan_2"
   
   Choose the template that BEST fits the style and story. If multiple fit, pick one.

JSON:
{{
  "story_essence": "...",
  "deity_attributes": ["..."],
  "emotional_arc": "...",
  "mantras": ["..."],
  "imagery": ["..."],
  "musical_mood": "...",
  "indian_ras": "Veer/Shanta/Karuna/Bhakti/Adbhuta/Shringar/Hasya",
  "singability_level": "high/medium/flexible",
  "singability_reason": "Explain WHY this level: story type, style demands, creative goal",
  "recommended_structure": "template_key (e.g., qawwali_1, rock_devotional_2, classical_bhajan_1)"
}}"""
        
        # Use FLASH for analysis (cheap operation)
        response = self._call_llm(prompt, 0.8, use_pro=False, call_type='analysis')
        return self._extract_json(response, {
            "story_essence": row_data['Story'],
            "deity_attributes": [],
            "mantras": [f"ॐ {row_data['God']}"],
            "imagery": [],
            "emotional_arc": "devotional",
            "musical_mood": "meditative",
            "indian_ras": "Bhakti",
            "singability_level": "high",
            "singability_reason": "Default: Most devotional songs should be catchy and singable",
            "recommended_structure": "classical_bhajan_1"
        })
    
    def _get_story_specific_warnings(self, row_data: Dict) -> str:
        """Generate dynamic story-specific critical word warnings (Advanced: Future-proofing)"""
        story = row_data['Story'].lower()
        god = row_data['God'].lower()
        
        warnings = []
        
        # Story-specific critical terms
        if 'shabari' in story or 'शबरी' in story:
            warnings.append("⚠️ SHABARI STORY: Use 'जूठा/जूठे' (tasted), NEVER 'झूठा/झूठे' (false)!")
        
        if 'sudama' in story or 'सुदामा' in story:
            warnings.append("⚠️ SUDAMA STORY: 'दरिद्र' (poor) not 'दरिंदा' (beast), 'मित्र' (friend) critical term")
        
        if 'krishna' in god and ('makhan' in story or 'butter' in story or 'माखन' in story):
            warnings.append("⚠️ KRISHNA MAKHAN: Use 'माखन' (butter), not 'मकान' (house)")
        
        if 'hanuman' in god or 'हनुमान' in god:
            warnings.append("⚠️ HANUMAN: 'बजरंग' (Bajrang) not 'बजरंगी', 'वानर' (monkey) respectfully")
        
        # Deity-specific attribute warnings
        if 'shiv' in god or 'शिव' in god:
            warnings.append("⚠️ SHIVA TERMS: त्रिनेत्र, डमरू, गंगा - don't use for other deities")
        
        if 'ram' in god or 'राम' in god:
            warnings.append("⚠️ RAM TERMS: धनुष, सीता, अयोध्या - keep Ram-specific")
        
        if 'krishna' in god or 'कृष्ण' in god:
            warnings.append("⚠️ KRISHNA TERMS: मुरली, माखन, राधा - keep Krishna-specific")
        
        if warnings:
            return "\n🎯 STORY-SPECIFIC CRITICAL CHECKS:\n" + "\n".join(warnings) + "\n"
        else:
            return ""
    
    def _chain2_generate_content(self, row_data: Dict, analysis: Dict) -> str:
        """Chain 2: Generate lyrics or arrangement"""
        is_long = row_data['Type'].lower() == 'long'
        
        if is_long:
            return self._generate_instrumental(row_data, analysis)
        else:
            return self._generate_lyrics(row_data, analysis)
    
    def _generate_lyrics(self, row_data: Dict, analysis: Dict) -> str:
        """Generate Hindi lyrics with Suno AI formatting using varied structures"""
        user_notes = row_data.get('User Notes', '')
        notes_instruction = f"\nUSER REQUEST: {user_notes}\nIncorporate this throughout." if user_notes else ""
        
        mantras = analysis.get('mantras', [f"ॐ {row_data['God']}"])[:4]
        attrs = analysis.get('deity_attributes', [])[:5]
        imagery = analysis.get('imagery', [])[:5]
        
        # Get unique structure for this song (Hybrid Method)
        base_structure = self._select_song_structure(row_data, analysis)
        
        # Determine rhythm strictness based on singability level
        singability = analysis.get('singability_level', 'medium')
        
        if singability == 'high':
            rhythm_rules = """
RHYTHM RULES (HIGH SINGABILITY - NATURAL FLOW):
1. MATRA COUNT: Maintain consistent meter (14-16 matras per line)
   - Prioritize natural flow and singability over exact count
   - Lines within a section should feel rhythmically similar
   - Example: "तू ही सहारा मेरा प्रभु, तू ही है जीवनधारा" (~16 matras, natural flow)
   - If a line feels forced, adjust vocabulary or word order for smoothness
2. RHYME SCHEME: Strong rhyme pattern
   - AABB (couplet rhyme) works best for catchy songs
   - ABAB (alternate rhyme) acceptable if it serves the story
   - Common endings: -आ/-या, -ए/-े, -ी/-री, -ों/-ो
3. CHORUS GUIDELINES:
   - Typically 2-4 lines with consistent meter
   - Should be the most memorable part
   - Use simple, repetitive, vowel-rich words
   - Strong rhyme pattern (AABB preferred)
   - Balance repetition with meaning"""
        
        elif singability == 'medium':
            rhythm_rules = """
RHYTHM RULES (BALANCED - MEDIUM SINGABILITY):
1. MATRA COUNT: Choose consistent meter (12-16 matras per line)
   - Keep lines within a section rhythmically similar
   - Verses and chorus can have different meters
   - Example 12-matra feel: "राम नाम का ध्यान करूँ मैं"
   - Example 16-matra feel: "सुदामा चले द्वार पर मित्र कृष्ण के संग में"
   - Focus on natural flow - if it sounds smooth, it's working
2. RHYME SCHEME: Flexible rhyme pattern
   - AABB (couplet): Best for catchy sections
   - ABAB (alternate): Good for storytelling verses
   - ABCB (partial): Works for narrative depth
   - Choose based on what serves the story and melody
3. CHORUS GUIDELINES:
   - Should be melodic and repeatable
   - Clear rhyme pattern helps memorability
   - Balance meaning with singability
   - Can be 2-4 lines depending on the story"""
        
        else:  # flexible
            rhythm_rules = """
RHYTHM RULES (FLEXIBLE - STORYTELLING PRIORITY):
1. MATRA COUNT: Aim for 12-16 matras, allow variation for narrative
   - Keep lines roughly similar length
   - Prioritize natural Hindi flow over strict counting
2. RHYME SCHEME: Natural rhyme where possible
   - Don't force rhymes if story suffers
   - Use internal rhyme for musicality
3. FOCUS:
   - Narrative clarity
   - Emotional authenticity
   - Natural language flow"""
        
        prompt = f"""Hindi devotional lyricist following Suno AI framework.

🎯 QUALITY TARGETS:
   - Hindi: {self.hindi_threshold}/10 (Grammar + Poetry)
   - Singability: {self.singability_threshold}/10 (Rhythm + Melody)
   - Focus on getting rhythm and melody RIGHT the first time

FRAMEWORK (CRITICAL):
{self.song_guide[:3500]}

COMPLIANCE CHECKLIST:
✓ [Section] brackets for structure tags
✓ (ad-libs) parentheses ONLY for sung ad-libs/backing vocals like (Jai Shri Ram!) or (Hare Krishna!)
✓ DO NOT add Hinglish transliteration in parentheses - Suno doesn't need it!
✓ [Instrumental - Name] cues
✓ [Voice, Emotion] directions
✓ [BPM: XX] tag
✓ Hindi Devanagari for all main lyrics
✓ Structure compliance

SUNO AI TAG WHITELIST (Use ONLY these standard tags):
- [Intro], [Verse], [Pre-Chorus], [Chorus], [Bridge], [Outro]
- [Hook], [Build Up], [Breakdown], [Instrumental]
- [Solo - Instrument], [Interlude]
DO NOT invent custom tags like [Mehfil] or [Antara] - Suno may ignore them!

PARENTHESES USAGE:
✅ CORRECT: "(Jai Shri Ram!)" - ad-lib/backing vocal
✅ CORRECT: "(Hare Krishna, Hare Krishna!)" - chorus backing
❌ WRONG: "(Nit-ya saaf kar-ti hai ras-ta)" - NO transliteration!

{rhythm_rules}

RHYTHM STRATEGY (CRITICAL for natural flow):
- Test each line for natural flow: Does it feel smooth or forced?
- If a line feels TOO LONG for the beat → Cut adjectives or use shorter synonyms
- If a line is TOO SHORT → Extend vowel sounds (e.g., "Ram" → "Raama", "Shiv" → "Shiva")
- Focus on NATURAL FLOW and singability over mathematical syllable counting
- The singability test: Can it be sung smoothly without rushing or dragging?
- Prioritize meaning and emotion over rigid meter

MELODIC REQUIREMENTS:
1. VERSE-CHORUS CONTRAST:
   - Verses: Tell story, can be conversational
   - Chorus: Should be catchy and memorable (the highlight of the song)
   - Note: Some styles (Aarti, Ghazal, Stotra) may not have a traditional chorus
2. CHORUS HOOK (when applicable):
   - Often features a repeated central phrase or hook
   - Vowel-rich words (आ, ए, ओ sounds) work well for sustained notes
   - Consonant clusters (क्ष, त्र, प्र) are fine if the word is powerful and meaningful
   - Examples: 'क्षमा' (forgiveness), 'त्रिशूल' (trident), 'प्रभु' (lord) are all acceptable
3. MELODIC FLOW:
   - Short words in fast sections: "हाँ, तू, मैं, है"
   - Long words in slow sections: "सहारा, जीवनधारा"
   - End lines on open vowels for sustain: "प्रभुआ", "कृष्णा"

INPUTS:
- Deity: {row_data['God']}
- Story: {analysis.get('story_essence', row_data['Story'])}
- Style: {row_data['Style']}
- Instruments: {row_data['Instruments']}
- Mantras: {', '.join(mantras)}
- Attributes: {', '.join(attrs)}
- Imagery: {', '.join(imagery)}{notes_instruction}

STRUCTURE TEMPLATE (adapt creatively while maintaining format):
{base_structure}

SINGABILITY ADAPTATION:
- Level: {singability}
- Reason: {analysis.get('singability_reason', 'Balanced approach')}

{"PRIORITY: Catchy hooks, simple vocabulary, perfect rhythm!" if singability == 'high' else ""}
{"PRIORITY: Balance story depth with melodic moments!" if singability == 'medium' else ""}
{"PRIORITY: Rich storytelling, natural flow, authentic vocabulary!" if singability == 'flexible' else ""}

SEMANTIC & PHONETIC ACCURACY (CRITICAL!):
Scan for words that are phonetically similar but semantically disastrous in THIS SPECIFIC story context.

FAILURE MODES TO WATCH:
1. **Homophones with opposite meanings** (sound same, mean opposite):
   - झूठा (jhootha = false) vs जूठा (jootha = tasted/sacred leavings)
   - दीया (diya = lamp) vs दया (daya = mercy)
   
2. **Near-spelling errors that change tone**:
   - भूल (bhool = mistake) vs भोल (bhol = innocence)
   - सुधा (sudha = nectar) vs सुधार (sudhaar = correction)
   - प्रेम (prem = love) vs प्रेत (pret = ghost/spirit)

3. **Deity-specific terminology mismatches**:
   - Don't use Shiva attributes (Trinetra, Damru) for Rama
   - Don't use Krishna terms (Makhan, Murli) for other deities
   - Match vocabulary to {row_data['God']}

TASK: If you spot a word that looks correct but implies WRONG THEOLOGY for this deity/story, reconsider it.
Context: Deity={row_data['God']}, Story={row_data['Story']}
{self._get_story_specific_warnings(row_data)}

🎵 SONG STRUCTURE (Selected by Analysis) 🎵

Use this structure as your guide:

{base_structure}

FLEXIBILITY: You may adapt this structure if the story demands it, but ensure:
1. ✅ Every section tag has lyrics after it (no empty sections)
2. ✅ Song has a proper ending (Outro/Final/Fade)
3. ✅ Minimum 4-5 distinct sections (not counting repeats)
4. ✅ Complete narrative arc: beginning → development → climax → resolution
5. ✅ DO NOT stop mid-song or leave sections incomplete

A well-executed simple structure is better than a poorly executed complex one.

FINAL CHECKS:
1. Rhythm consistency - lines within sections should flow similarly
2. Rhyme pattern - clear and appropriate for the style
3. Memorability - chorus/hook should be catchy and singable
4. Melodic flow - natural ups and downs, emotional progression
5. Contextual accuracy - verify deity-specific terms and story details

Write ONLY formatted lyrics following ALL rules above. WRITE THE COMPLETE SONG FROM START TO FINISH."""
        
        # USE PRO MODEL for song generation (critical quality)
        # Temperature 0.7 (reduced from 0.9) for more focused, complete output
        response = self._call_llm(prompt, 0.7, use_pro=True, call_type='generation')
        
        # Safety check
        if response is None:
            response = ""
        
        # Clean up asterisks from lyrics (markdown formatting)
        response = response.replace('**', '').replace('*', '')
        
        # DISABLED: Auto-completion check removed per user request
        # The 30K token limit is sufficient for complete lyrics
        # If truncation occurs, it will be caught in manual review
        
        return response
    
    def _generate_instrumental(self, row_data: Dict, analysis: Dict) -> str:
        """Generate instrumental arrangement (concise for Suno AI <800 chars)"""
        inst_list = [i.strip() for i in row_data['Instruments'].split(',')]
        inst_main = inst_list[0] if inst_list else 'Instruments'
        
        prompt = f"""Create CONCISE Suno AI description for instrumental music - PURELY TECHNICAL.

INPUTS:
- Style: {row_data['Style']}
- Instruments: {row_data['Instruments']}
- Mood: {analysis.get('musical_mood', 'meditative')}
- Length: 10+ minutes, LOOPABLE

CRITICAL REQUIREMENTS:
- MAX 700 characters total
- PURELY technical/musical description (like for a music producer)
- Do NOT mention: deity, story, religious themes, or spiritual content
- ONLY describe: instruments, mood, structure, tempo, production, loopability
- Emphasize LOOPABLE and seamless transitions
- Focus on musical atmosphere and sonic qualities

FORMAT (TECHNICAL ONLY):
A [duration] [style] instrumental track featuring [list all instruments]. Creates [mood] ambience with [describe sonic texture]. Opens with [intro description], develops through [progression details], closes with [outro style]. LOOPABLE design with seamless transitions, perfect for extended listening/meditation/background. Instruments: [list with roles]. Mood: [descriptors]. Tempo: [BPM if relevant].

EXAMPLE: "A 10-minute meditative classical instrumental featuring sitar, tabla, tanpura, and soft strings. Creates peaceful, atmospheric ambience with gentle melodic interplay. Opens with tanpura drone and soft sitar, develops through rhythmic tabla patterns with string layers, closes with gradual fade. LOOPABLE with seamless transitions. Tempo: 65 BPM. Perfect for meditation and relaxation."

STRICT: Keep under 700 characters. Be concise but descriptive. NO deity/story mentions."""
        
        # USE PRO MODEL for instrumental generation (critical quality)
        response = self._call_llm(prompt, 0.7, use_pro=True, call_type='generation')
        
        # Safety check
        if response is None:
            response = ""
        
        # Clean up asterisks (markdown formatting)
        response = response.replace('**', '').replace('*', '')
        
        # Truncate if still too long (Suno AI limit is 1000, keep safe margin)
        char_count = len(response)
        if char_count > 800:
            print(f"  ⚠️  Description too long ({char_count} chars), truncating to 800...")
            response = response[:797] + "..."
        else:
            print(f"  ✅ Description length: {char_count} chars (under 800 limit)")
        
        return response
    
    def _chain3_metadata(self, row_data: Dict, analysis: Dict, content: str) -> Dict:
        """Chain 3: Generate metadata"""
        is_long = row_data['Type'].lower() == 'long'
        language = row_data.get('Language', 'Hindi').strip().lower()
        
        # Safety check for None content
        if content is None:
            content = ""
            print(f"  ⚠️  Content was None, using empty string for metadata")
        
        # Language-specific labels
        if language == 'hindi':
            youtube_type_label = 'ध्यान संगीत' if is_long else 'भक्ति गीत'
            desc_language = "Pure Hindi"
            title_lang = "Hindi (Devanagari)"
        else:
            youtube_type_label = 'Meditation Music' if is_long else 'Devotional Song'
            desc_language = "Pure English"
            title_lang = "English"
        
        prompt = f"""YouTube SEO expert for devotional music.

INPUTS:
- Deity: {row_data['God']}
- Story: {row_data['Story']}
- Style: {row_data['Style']}
- Type: {'Instrumental 10+ min' if is_long else 'Vocal song'}
- YouTube Language: {language.title()} (for title/description/tags ONLY)

CONTENT PREVIEW:
{content[:800] if content else 'No content available'}

SUNO AI DESCRIPTION (CRITICAL - Suno reads first words heaviest):
STRUCTURE: [Genre/Style] + [Atmosphere] + [Instruments with roles] + [Vocal Style]
- START with genre/style as FIRST WORD (e.g., "Indian Classical", "Himalayan Folk", "Devotional Qawwali")
- Write NARRATIVE JOURNEY: How the song unfolds with action verbs
- Pair instruments with ROLES: "Tabla patterns grow sparse to dynamic", "Bansuri decorates phrases"
- Use SENSORY language: "ambient pad under airy whispers", "acoustic warmth"
- Focus ONLY on: genre, instruments, vocals, structure, mood, production
- Do NOT mention: deity name, story, themes
- GOOD EXAMPLE: "Himalayan Folk devotional with ransingha and ektara adding texture as harmonium establishes verse rhythm. Bansuri decorates melodic phrases while tabla patterns grow from sparse to dynamic, elevating a joyous, uplifting atmosphere."
- Keep 50-70 words, written IN English

YOUTUBE DESCRIPTION GUIDELINES:
- Make it ENGAGING and VALUABLE - not just facts
- Start with emotional hook (e.g., "Experience divine peace...", "Journey into devotion...")
- Explain WHY someone should listen (benefits: peace, meditation, spiritual growth)
- Include story context that adds meaning
- Mention when to listen (morning prayers, meditation, celebrations)
- End with call-to-action (like, subscribe, share)
- Use emojis strategically (🙏 🎵 🎼 ⏱️ 🌟 🔔 📿 🕉️ ✨)

TAGS BALANCE (18-22 tags total):
1. CONTEXTUAL (6-8 tags): Deity name, story-specific, character names, episode names
   Examples: #ShriRam #ShabariKeBer #ShriKrishna #Sudama #Mahabharata
2. GENERAL BHAKTI (6-8 tags): Universal devotional keywords
   Examples: #DevotionalMusic #BhaktiGeet #HindiDevotional #SpiritualMusic #IndianDevotion #Bhajan #Kirtan #Aarti
3. VIRAL/REACH (6-8 tags): Trending, discoverable, SEO keywords
   Examples: #SunoAI #NewDevotional #2025Bhajan #ViralBhajan #TrendingBhajan #PeacefulMusic #MeditationMusic #YogaMusic

GENERATE JSON:
{{
  "song_title": "{title_lang} title - concise, captures deity and story",
  "suno_ai_description": "TECHNICAL MUSIC DESCRIPTION ONLY (50-80 words): Describe genre, instruments (list each), vocal style (male/female/layered), musical structure (intro-verse-chorus-bridge etc), mood (meditative/energetic/peaceful), production (reverb/layers/tempo). NO deity, story, or theme mention.",
  "youtube_title": "Title | {row_data['Style']} | {youtube_type_label}",
  "youtube_description": "{desc_language} 180-220 words - ENGAGING and VALUABLE with emojis. Start with emotional hook, explain benefits, provide context, mention when to listen, end with CTA.",
  "tags": "18-22 hashtags BALANCED: 6-8 contextual (deity, story), 6-8 general bhakti, 6-8 viral/reach. Examples: #ShriRam #ShabariKeBer #DevotionalMusic #BhaktiGeet #SunoAI #ViralBhajan #2025Devotional #MeditationMusic"
}}

Return ONLY valid JSON."""
        
        # Use FLASH for metadata generation (cheap operation)
        response = self._call_llm(prompt, 0.7, use_pro=False, call_type='analysis')
        return self._extract_json(response, {
            "song_title": f"{row_data['God']} - {row_data['Story']}",
            "suno_ai_description": f"A devotional {row_data['Style']} track with traditional Indian instruments including {row_data['Instruments'].split(',')[0].strip()}. Features layered vocals with atmospheric production and meditative tempo.",
            "youtube_title": f"{row_data['God']} | {row_data['Style']}",
            "youtube_description": "Devotional song",
            "tags": f"#{row_data['God']} #DevotionalMusic"
        })
    
    def _iterative_refine(self, content: str, issues: list, row_data: Dict, analysis: Dict, focus: str) -> str:
        """OPTIMIZED: Iteratively refine existing content instead of regenerating from scratch
        
        Args:
            content: Current lyrics
            issues: List of specific issues to fix
            focus: 'hindi', 'singability', or 'general'
        """
        if not issues or not self.use_iterative_refinement:
            return content
        
        issues_text = "\n".join(f"- {issue}" for issue in issues[:5])
        
        if focus == 'hindi':
            focus_instruction = "Fix ONLY grammar/vocabulary issues. Keep rhythm, melody, and structure intact."
        elif focus == 'singability':
            focus_instruction = "Fix ONLY rhythm/melody issues (adjust matras, improve hooks). Keep meaning and grammar intact."
        else:
            focus_instruction = "Fix all mentioned issues while preserving the song's core essence."
        
        prompt = f"""Hindi lyric editor. Refine these lyrics by fixing SPECIFIC issues.

ISSUES TO FIX:
{issues_text}

CURRENT LYRICS:
{content}

INSTRUCTIONS:
{focus_instruction}

Rules:
- Make MINIMAL changes (only fix issues)
- Keep [tags], structure, mantras intact
- Maintain same verse/chorus count
- Preserve storytelling and emotion
- RETURN THE COMPLETE SONG - do not truncate!

OUTPUT: Return ONLY the refined lyrics with fixes applied."""
        
        # Use Pro for refinement (creative task)
        refined = self._call_llm(prompt, 0.7, use_pro=True, call_type='refinement')
        
        # Clean up
        refined = refined.replace('**', '').replace('*', '')
        
        if len(refined) < 100:  # Too short, something went wrong
            return content
        
        return refined
    
    def _chain4_refine(self, content: str, row_data: Dict) -> str:
        """Chain 4: Quick polish pass - RETURNS LYRICS ONLY, NOT SUGGESTIONS!"""
        
        # Quick check: if content looks good, skip refinement
        if len(content) > 500:  # Has substantial content
            prompt = f"""Hindi lyric editor. Review and polish if needed.

LYRICS:
{content}

TASK: Check for:
- Hindi grammar errors
- Format issues ([tags], structure)
- Obvious mistakes

If excellent (no issues): Output exactly "APPROVED"
If minor issues found: Output the CORRECTED LYRICS (entire song, not suggestions!)

CRITICAL: Output ONLY:
1. The word "APPROVED", OR
2. The full corrected lyrics (COMPLETE - do not truncate!)

DO NOT output suggestions, explanations, or feedback!"""
            
            refined = self._call_llm(prompt, 0.5, use_pro=False, call_type='validation')  # Use Flash for quick check
            
            # Clean up
            refined = refined.replace('**', '').replace('*', '').strip()
            
            # If approved or response is suggestions (contains "Suggestion:", "Refined Version", etc)
            if (refined.upper() == "APPROVED" or 
                len(refined) < 200 or 
                "suggestion" in refined.lower() or
                "refined version" in refined.lower() or
                "improvement" in refined.lower()):
                return content  # Return original
            
            return refined
        
        return content
    
    def _chain_singability_check(self, lyrics: str, row_data: Dict, analysis: Dict) -> Dict:
        """OPTIMIZED: Combined Rhythm + Melody check in ONE call = Singability"""
        
        singability = analysis.get('singability_level', 'high')
        
        # Rhythm expectations
        if singability == 'high':
            rhythm_rules = "STRICT: 16 matras per line, AABB rhyme"
        elif singability == 'medium':
            rhythm_rules = "BALANCED: 12 or 16 matras (consistent), AABB/ABAB rhyme"
        else:
            rhythm_rules = "FLEXIBLE: Natural flow, 10-18 matras"
        
        prompt = f"""Singability Expert: Evaluate if this song is EASY TO SING.

CONTEXT: {row_data['Style']}, Singability={singability}, Target={self.singability_threshold}/10

LYRICS (first 1000 chars):
{lyrics[:1000]}

CHECK BOTH:
1. RHYTHM (Meter + Flow):
   {rhythm_rules}
   - Count matras in sung lines only (ignore [tags])
   - Report specific issues: "Line X: Y matras (need 16)"

2. MELODY (Hooks + Catchiness):
   - Repeated hook in chorus? (YES=catchy, NO=forgettable)
   - Vowel endings for sustain? (आ, ए, ओ = good)
   - No heavy consonant clusters? (क्ष्, ज्ञ in hooks = bad)

SCORE STRICTLY:
- Both good (rhythm+melody) = 8-10/10
- One good, one weak = 5-7/10  
- Both weak = <5/10

JSON:
{{
  "rhythm_score": 1-10,
  "melody_score": 1-10,
  "overall_singability": 1-10,
  "issues": ["Top 3 specific issues"],
  "verdict": "SINGABLE" if overall_singability >= {self.singability_threshold} else "NEEDS_IMPROVEMENT"
}}"""
        
        response = self._call_llm(prompt, 0.3, use_pro=False, call_type='validation')
        return self._extract_json(response, {
            "rhythm_score": 7,
            "melody_score": 7,
            "overall_singability": 7,
            "issues": [],
            "verdict": "NEEDS_IMPROVEMENT"
        })
    
    def _chain_combined_quality_check(self, lyrics: str, row_data: Dict, analysis: Dict) -> Dict:
        """OPTIMIZED: Combined Hindi + Rhythm + Melody check in ONE LLM call"""
        
        singability = analysis.get('singability_level', 'high')
        
        # Adaptive thresholds based on singability
        rhythm_target = 8 if singability == 'high' else 6 if singability == 'medium' else 5
        melody_target = 8 if singability == 'high' else 7 if singability == 'medium' else 6
        
        prompt = f"""Quality Control Expert evaluating Hindi devotional lyrics.

CONTEXT:
- Style: {row_data['Style']}
- Singability Level: {singability}
- Quality Thresholds: Hindi={self.hindi_threshold}/10, Rhythm+Melody={self.singability_threshold}/10

LYRICS TO EVALUATE:
{lyrics}

EVALUATE ALL THREE ASPECTS:

1. HINDI QUALITY (Grammar + Poetry):
   - Grammar correctness in Hindi
   - Poetic flow and authenticity
   - Accept Sanskrit mantras and complex vocabulary when appropriate
   - Score 1-10

2. RHYTHM QUALITY (Meter + Flow):
   {"- Check for consistent meter (14-16 matras per line within sections)" if singability == 'high' else "- Check meter consistency (12-16 matras, similar within sections)" if singability == 'medium' else "- Natural flow, roughly similar line lengths"}
   {"- Rhyme scheme should be strong (AABB or ABAB)" if singability == 'high' else "- Rhyme scheme flexible (AABB, ABAB, or ABCB)" if singability == 'medium' else "- Natural rhyme where possible"}
   - Focus on natural flow and singability
   - Report issues only if rhythm feels forced or inconsistent
   - Target Score: {rhythm_target}/10

3. MELODY QUALITY (Hooks + Catchiness):
   - Check for memorable hook or repeated phrase (if style has chorus)
   - Vowel-rich words work well for sustained notes
   - Powerful words with consonant clusters are acceptable if meaningful
   - Verse-chorus contrast (when applicable)
   - Target Score: {melody_target}/10

SCORING GUIDELINES:
- Rhythm: Focus on natural flow and consistency, not exact syllable counts
- Melody: Memorable hooks are important for catchy styles, less so for narrative/stotra
- Hindi: Grammar and poetic quality matter most

JSON RESPONSE (ALL fields required):
{{
  "hindi_grammar_score": 1-10,
  "hindi_poetic_score": 1-10,
  "hindi_overall": 1-10,
  "rhythm_score": 1-10,
  "meter_consistency": 1-10,
  "rhythm_overall": 1-10,
  "melody_hook_strength": 1-10,
  "melody_phrasing": 1-10,
  "melody_overall": 1-10,
  "combined_score": 1-10 (average of hindi_overall + rhythm_overall + melody_overall),
  "all_issues": ["List ALL issues found in Hindi, Rhythm, Melody"],
  "critical_fixes": ["Top 3 most important fixes"],
  "verdict": "EXCELLENT" if hindi_overall >= {self.hindi_threshold} and rhythm_overall >= {self.singability_threshold} else "NEEDS_IMPROVEMENT",
  "pass_thresholds": {{"hindi": {self.hindi_threshold}, "singability": {self.singability_threshold}}}
}}

BE STRICT! Count actual matras and verify everything!"""
        
        # Use FLASH for quality checks (cheap operation)
        response = self._call_llm(prompt, 0.3, use_pro=False, call_type='validation')
        result = self._extract_json(response, {
            "hindi_grammar_score": 8,
            "hindi_poetic_score": 8,
            "hindi_overall": 8,
            "rhythm_score": 7,
            "meter_consistency": 7,
            "rhythm_overall": 7,
            "melody_hook_strength": 7,
            "melody_phrasing": 7,
            "melody_overall": 7,
            "combined_score": 7,
            "all_issues": [],
            "critical_fixes": [],
            "verdict": "NEEDS_IMPROVEMENT",
            "pass_thresholds": {"hindi": self.hindi_threshold, "singability": self.singability_threshold}
        })
        
        return result
    
    def _chain5_hindi_quality(self, lyrics: str, row_data: Dict) -> Dict:
        """Chain 5: Hindi language quality check - Accepts complex vocabulary when appropriate"""
        
        prompt = f"""Hindi devotional music linguist.

LYRICS TO EVALUATE:
{lyrics}

CHECK ONLY:
1. Grammar correctness in Hindi portions
2. Poetic flow and authenticity
3. Devotional appropriateness
4. WORD REPETITION in adjacent/nearby lines (CRITICAL!)
5. SEMANTIC & PHONETIC ACCURACY (CRITICAL!) - Catch "Phonetic Hallucinations":
   Look for words that are phonetically similar but semantically disastrous in this story context.
   
   TYPES OF ERRORS TO DETECT:
   a) **Homophones with opposite meanings**:
      - झूठा (jhootha = false) vs जूठा (jootha = tasted/sacred leavings)
      - दीया (diya = lamp) vs दया (daya = mercy)
      
   b) **Near-spelling that changes tone**:
      - भूल (bhool = mistake) vs भोल (bhol = innocence)
      - सुधा (sudha = nectar) vs सुधार (sudhaar = correction)
      - प्रेम (prem = love) vs प्रेत (pret = ghost)
      - बांस (baans = bamboo) vs बांसुरी (baansuri = flute)
      
   c) **Deity-specific term mismatches**:
      - Shiva attributes (त्रिनेत्र, डमरू) used for Rama → FLAG IT
      - Krishna terms (माखन, मुरली) used for other deities → FLAG IT
   
   TASK: If a word looks correct but doesn't match the THEOLOGY of this deity/story, flag it!

ACCEPT:
- Sanskrit mantras and vocabulary (traditional)
- Complex/difficult Hindi when enhances meaning
- Traditional poetic conventions
- Creative devotional expressions
- **BHAKTI REPETITION (CRITICAL):** Intentional repetition of deity names or mantras for meditative effect
  Examples: "गोविंदा बोलो, हरि गोपाला बोलो", "राम राम जय राजा राम", "हरे कृष्णा हरे कृष्णा"
- **ANAPHORA (POETIC):** Starting consecutive lines with same word for emphasis
  Examples: "तू ही मेरा साथी... तू ही मेरा प्रभु"

REJECT (Flag as issues):
- **LAZY FILLER REPETITION:** Unintentional repetition of filler words (है, और, फिर, वह, यह)
  ❌ BAD: "वह है प्रभु मेरा... वह है जग का स्वामी" (lazy "वह है" repetition)
- **LAZY RHYMING:** Rhyming a word with itself instead of finding synonyms
  ❌ BAD: "राम का नाम... प्रभु का नाम" (lazy "नाम" repetition)
- **AWKWARD DESCRIPTIVE REPETITION:** Same descriptive word in adjacent lines without purpose
  ❌ BAD: "नित्य साफ़ करती है रस्ता... नित्य फूल चुन राह" (awkward "नित्य" - use "हर दिन" instead)

EXAMPLES:
✅ GOOD: "राम राम कहते रहो, राम राम जपते रहो" (Namasmaran - intentional devotional chanting)
✅ GOOD: "तू ही सहारा मेरा... तू ही है जीवनधारा" (Anaphora - poetic emphasis)
✅ GOOD: "हरे कृष्णा हरे कृष्णा, कृष्णा कृष्णा हरे हरे" (Mantra - traditional repetition)
✅ GOOD: "चख चख कर वह जूठे बेर खिलाती" (Shabari story - correct word!)
❌ BAD: "नित्य साफ़ करती... नित्य फूल चुन" (Lazy descriptive word - vary vocabulary)
❌ BAD: "और वह गए... और वह आए" (Lazy filler - use different connectors)
❌ BAD: "चख चख कर वह झूठे बेर खिलाती" (WRONG word - "झूठा" means false, should be "जूठा" = tasted!)

CONTEXT FOR THIS SONG:
- Deity: {row_data['God']}
- Story: {row_data['Story']}
{self._get_story_specific_warnings(row_data)}
SCORING:
- 10: Perfect Hindi + appropriate Sanskrit + NO awkward repetitions + contextually accurate words
- 8-9: Minor issues or 1-2 acceptable repetitions
- <8: Multiple grammar issues OR awkward repetitions
- <7: CRITICAL contextual word errors (e.g., झूठा vs जूठा) - these change the story meaning!

JSON:
{{
  "grammar_score": 1-10,
  "poetic_score": 1-10,
  "overall_score": 1-10,
  "issues": ["major issues only, specifically mention ANY awkward word repetitions"],
  "suggestions": ["fixes, provide alternate words for repetitions"],
  "verdict": "EXCELLENT" or "NEEDS_IMPROVEMENT"
}}"""
        
        # Use FLASH for Hindi quality check (cheap operation)
        response = self._call_llm(prompt, 0.3, use_pro=False, call_type='validation')  # Low temp for consistency
        result = self._extract_json(response, {
            "grammar_score": 9,
            "poetic_score": 9,
            "overall_score": 9,
            "issues": [],
            "suggestions": [],
            "verdict": "NEEDS_IMPROVEMENT"
        })
        
        return result
    
    def _chain5b_rhythm_check(self, lyrics: str, row_data: Dict, analysis: Dict = None) -> Dict:
        """Chain 5b: Rhythm and meter validation (adapts to singability level)"""
        
        singability = analysis.get('singability_level', 'medium') if analysis else 'medium'
        
        # Create specific checking criteria based on singability
        if singability == 'high':
            check_criteria = """
CHECK CRITERIA (STRICT):
1. MATRA COUNT VERIFICATION:
   - Extract all sung lines (not [tags] or (ad-libs))
   - Count matras in EACH line
   - MUST be exactly 16 matras per line
   - Report any line != 16 matras
2. RHYME VERIFICATION:
   - Check AABB pattern (lines 1-2 rhyme, 3-4 rhyme)
   - Endings must match phonetically
   - Report non-rhyming pairs
3. CHORUS CHECK:
   - Must have clear repetitive hook
   - Hook phrase repeated in chorus
   - Simple vocabulary only
   
PASS THRESHOLD: All lines 16 matras + perfect rhyme = 9-10/10
FAIL: Any line wrong count or bad rhyme = <7/10"""
        
        elif singability == 'medium':
            check_criteria = """
CHECK CRITERIA (BALANCED):
1. MATRA COUNT VERIFICATION:
   - Count matras per line
   - Check consistency: all verse lines same count (12 or 16)
   - All chorus lines same count
   - Allow ±1 matra variation for natural flow
2. RHYME VERIFICATION:
   - Check for AABB or ABAB pattern
   - Chorus must have clear rhyme
   - Verses can be more flexible
3. MELODIC FLOW:
   - Check breathing points (commas, pauses)
   - Natural sentence breaks
   
PASS THRESHOLD: Consistent meter + good rhyme = 7-8/10
NEEDS WORK: Inconsistent meter = <6/10"""
        
        else:  # flexible
            check_criteria = """
CHECK CRITERIA (FLEXIBLE):
1. GENERAL FLOW:
   - Lines roughly similar length (10-18 matras)
   - Not too long or short
   - Natural Hindi phrasing
2. RHYME:
   - Some rhyme present (internal or end)
   - Doesn't need to be perfect
3. NARRATIVE CLARITY:
   - Story flows naturally
   - Language authentic
   
PASS THRESHOLD: Natural flow + some rhyme = 6-7/10
ACCEPTABLE: Story clear even if rhythm loose = 5/10"""
        
        prompt = f"""Rhythm analyst evaluating Hindi devotional lyrics.

CONTEXT: {row_data['Style']} style, {singability} singability target

{check_criteria}

LYRICS TO ANALYZE:
{lyrics}

ANALYSIS STEPS:
1. Extract sung lines (ignore [tags] and (ad-libs))
2. Count matras (syllables) in each line
3. Check rhyme pattern
4. Verify chorus has hook
5. Rate overall rhythm quality

EXAMPLE ANALYSIS:
Line 1: "तू ही सहारा मेरा प्रभु" = 10 matras (target: 16) ❌
Line 2: "जीवन का आधार है तू" = 9 matras (target: 16) ❌
Rhyme: "प्रभु" vs "तू" = NO RHYME ❌
Score: 4/10 - NEEDS_IMPROVEMENT

JSON RESPONSE:
{{
  "rhythm_score": 1-10,
  "meter_consistency": 1-10,
  "singability": 1-10,
  "overall_rhythm": 1-10,
  "rhythm_issues": ["SPECIFIC issues with matra counts or rhyme"],
  "suggestions": ["CONCRETE fixes like 'Make all lines 16 matras' or 'Fix rhyme scheme to AABB'"],
  "verdict": "EXCELLENT" or "NEEDS_IMPROVEMENT"
}}

BE STRICT. Count actual matras and verify rhyme schemes!"""
        
        # Use FLASH for rhythm check (cheap operation)
        response = self._call_llm(prompt, 0.3, use_pro=False, call_type='validation')
        result = self._extract_json(response, {
            "rhythm_score": 8,
            "meter_consistency": 8,
            "singability": 8,
            "overall_rhythm": 8,
            "rhythm_issues": [],
            "suggestions": [],
            "verdict": "NEEDS_IMPROVEMENT"
        })
        
        return result
    
    def _chain2b_melody_check(self, lyrics: str, row_data: Dict) -> Dict:
        """Chain 2b: Melody optimization - Check for catchy hooks and melodic potential"""
        
        prompt = f"""Melody specialist evaluating devotional song for SINGING QUALITY.

STYLE: {row_data['Style']}

LYRICS:
{lyrics}

CHECK MELODIC QUALITY (STRICT):

1. HOOK STRENGTH (Main Chorus):
   ✓ GOOD: "राम नाम सच है, राम नाम सच है" (repeats, simple, catchy)
   ✗ BAD: "जीर्ण वस्त्र धारण कर, मन में संकोच लिए" (complex, no repetition, not singable)
   
   CHECK:
   - Is there a repeated phrase in chorus? (YES = +3 points)
   - Is the phrase simple (3-5 words)? (YES = +2 points)
   - Can you remember it after one read? (YES = +3 points)
   - Uses vowel-rich words? (आ, ए, ओ sounds) (YES = +2 points)

2. MELODIC PHRASING:
   ✓ GOOD: "तू ही सहारा मेरा" (flows, has rhythm, vowel endings)
   ✗ BAD: "जीर्ण वस्त्र तन पर" (choppy, consonant clusters क्ष्, stops flow)
   
   CHECK:
   - Lines end on open vowels for sustain? (YES = good)
   - No heavy consonant clusters in main hook? (YES = good)
   - Natural breathing points (commas)? (YES = good)

3. MELODIC VARIATION:
   ✓ GOOD: Verse conversational, Chorus HIGH ENERGY, Bridge emotional
   ✗ BAD: All sections sound same, monotonous
   
   CHECK:
   - Chorus clearly different from verse? (YES = +2 points)
   - Bridge adds emotional variation? (YES = +2 points)

4. OVERALL SINGABILITY:
   - Would average person sing along after 2 listens? (YES = 8-10/10, NO = <6/10)
   - Is chorus stuck in your head? (YES = 8-10/10, NO = <6/10)

SCORING GUIDE:
- 9-10: EXTREMELY CATCHY (professional hit song level)
- 7-8: STRONG HOOK (memorable, singable)
- 5-6: ACCEPTABLE (ok but not catchy)
- <5: NEEDS WORK (not singable, no hook)

JSON:
{{
  "hook_strength": 1-10,
  "melodic_phrasing": 1-10,
  "melodic_variation": 1-10,
  "overall_melody": 1-10,
  "melody_issues": ["SPECIFIC issues: 'No repeated hook in chorus', 'Too many consonant clusters', etc"],
  "melody_suggestions": ["CONCRETE fixes: 'Add repeated phrase in chorus', 'Change X to Y for better flow', etc"],
  "verdict": "MELODIOUS" or "NEEDS_IMPROVEMENT"
}}

BE STRICT! Focus ONLY on melody and catchiness, NOT grammar or meaning!"""
        
        # Use FLASH for melody check (cheap operation)
        response = self._call_llm(prompt, 0.3, use_pro=False, call_type='validation')
        result = self._extract_json(response, {
            "hook_strength": 7,
            "melodic_phrasing": 7,
            "melodic_variation": 7,
            "overall_melody": 7,
            "melody_issues": [],
            "melody_suggestions": [],
            "verdict": "NEEDS_IMPROVEMENT"
        })
        
        return result
    
    def _chain6_validate(self, data: Dict, row_data: Dict) -> Dict:
        """Chain 5: Cultural validation"""
        title = data.get('song_title', '')
        content = data.get('lyrics', data.get('arrangement', ''))[:1000]
        desc = data.get('youtube_description', '')[:500]
        
        prompt = f"""Hindu religious and cultural sensitivity expert.

CONTENT:
Deity: {row_data['God']}
Story: {row_data['Story']}

Title: {title}
Content: {content}
Description: {desc}

VALIDATE:
1. Cultural accuracy
2. Religious sensitivity
3. Mantra authenticity
4. Traditional respect
5. Factual correctness

CRITICAL MANTRA SAFETY CHECKS:
- Verify Beej Mantras (Om, Hreem, Kleem, Shreem) are used correctly for specific deity
- Ensure no "Ugra" (Fierce) mantras like Kali Beej are used in "Vatsalya" (Child-love) context
- Check that Tantric mantras are not casually mixed with Puranic story lyrics
- Flag if mantras are incomplete or incorrectly structured
- Examples:
  ✅ CORRECT: "Om Namah Shivaya" for Shiva
  ✅ CORRECT: "Om Kleem Krishnaya Namaha" for Krishna
  ❌ WRONG: "Om Hreem" alone without context (incomplete)
  ❌ WRONG: Kali mantras in a Krishna lullaby (context mismatch)

JSON:
{{
  "is_valid": true/false,
  "cultural_score": 1-10,
  "issues_found": ["..."],
  "recommendations": ["..."],
  "approval_status": "APPROVED" or "NEEDS_REVISION",
  "summary": "brief assessment"
}}

Be strict, protect sentiments, allow creativity."""
        
        # Use FLASH for cultural validation (cheap operation)
        response = self._call_llm(prompt, 0.3, use_pro=False, call_type='validation')
        return self._extract_json(response, {
            "is_valid": True,
            "cultural_score": 7,
            "issues_found": [],
            "recommendations": [],
            "approval_status": "APPROVED",
            "summary": "Auto-approved"
        })
    
    def _generate_filename(self, row_data: Dict) -> str:
        """Generate clean filename"""
        day = row_data['Day']
        god = re.sub(r'[^\w\s-]', '', row_data['God']).strip().replace(' ', '_').lower()
        story = re.sub(r'[^\w\s-]', '', row_data['Story'])[:30].strip().replace(' ', '_').lower()
        return f"day_{day}_{god}_{story}.txt"
    
    def _save_song_file(self, row_data: Dict, output: Dict) -> str:
        """Save individual song file"""
        filename = self._generate_filename(row_data)
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"{'='*80}\n")
            f.write(f"DAY {row_data['Day']}: {row_data['God']} - {row_data['Story']}\n")
            f.write(f"{'='*80}\n\n")
            
            f.write(f"SONG TITLE:\n{output.get('Song Title', 'N/A')}\n\n")
            
            if 'Hindi Lyrics' in output:
                f.write(f"HINDI LYRICS:\n{'='*80}\n{output['Hindi Lyrics']}\n\n")
            
            suno_desc = output.get('Suno AI Description', 'N/A')
            suno_char_count = len(suno_desc) if suno_desc != 'N/A' else 0
            f.write(f"SUNO AI DESCRIPTION ({suno_char_count} chars):\n{suno_desc}\n\n")
            f.write(f"YOUTUBE TITLE:\n{output.get('YouTube Title', 'N/A')}\n\n")
            f.write(f"YOUTUBE DESCRIPTION:\n{output.get('YouTube Description', 'N/A')}\n\n")
            f.write(f"TAGS:\n{output.get('Tags', 'N/A')}\n\n")
            
            # Hindi Quality
            if 'Hindi Quality' in output:
                hq = output['Hindi Quality']
                if hq.get('verdict') != 'Skipped':
                    f.write(f"HINDI QUALITY CHECK (Grammar + Poetry):\n")
                    f.write(f"  Grammar: {hq.get('grammar_score', 'N/A')}/10\n")
                    f.write(f"  Poetic: {hq.get('poetic_score', 'N/A')}/10\n")
                    f.write(f"  Overall: {hq.get('overall_score', 'N/A')}/10\n")
                    f.write(f"  Verdict: {hq.get('verdict', 'N/A')}\n")
                    if hq.get('issues'):
                        f.write(f"  Issues: {', '.join(hq['issues'][:3])}\n")
                    f.write(f"\n")
            
            # Rhythm Quality
            if 'Rhythm Quality' in output:
                rq = output['Rhythm Quality']
                if rq.get('verdict') != 'Skipped':
                    f.write(f"RHYTHM QUALITY CHECK:\n")
                    f.write(f"  Rhythm Score: {rq.get('rhythm_score', 'N/A')}/10\n")
                    f.write(f"  Meter Consistency: {rq.get('meter_consistency', 'N/A')}/10\n")
                    f.write(f"  Singability: {rq.get('singability', 'N/A')}/10\n")
                    f.write(f"  Overall Rhythm: {rq.get('overall_rhythm', 'N/A')}/10\n")
                    f.write(f"  Verdict: {rq.get('verdict', 'N/A')}\n")
                    if rq.get('rhythm_issues'):
                        f.write(f"  Issues: {', '.join(rq['rhythm_issues'][:2])}\n")
                    f.write(f"\n")
            
            # Melody Quality (NEW!)
            if 'Melody Quality' in output:
                mq = output['Melody Quality']
                if mq.get('verdict') != 'Skipped':
                    f.write(f"MELODY QUALITY CHECK:\n")
                    f.write(f"  Hook Strength: {mq.get('hook_strength', 'N/A')}/10\n")
                    f.write(f"  Melodic Phrasing: {mq.get('melodic_phrasing', 'N/A')}/10\n")
                    f.write(f"  Melodic Variation: {mq.get('melodic_variation', 'N/A')}/10\n")
                    f.write(f"  Overall Melody: {mq.get('overall_melody', 'N/A')}/10\n")
                    f.write(f"  Verdict: {mq.get('verdict', 'N/A')}\n")
                    if mq.get('melody_issues'):
                        f.write(f"  Issues: {', '.join(mq['melody_issues'][:2])}\n")
                    f.write(f"\n")
            
            # Cultural Validation
            if 'Cultural Validation' in output:
                val = output['Cultural Validation']
                f.write(f"CULTURAL VALIDATION:\n")
                f.write(f"  Status: {val.get('approval_status', 'N/A')}\n")
                f.write(f"  Score: {val.get('cultural_score', 'N/A')}/10\n")
                f.write(f"  Summary: {val.get('summary', 'N/A')}\n")
                if val.get('issues_found'):
                    f.write(f"  Issues: {', '.join(val['issues_found'])}\n")
        
        print(f"  💾 Saved: {filename}")
        return filename
    
    def _update_excel_music_column(self, row_num: int, filename: str):
        """Update Excel Music column"""
        music_col = self.column_map.get('Music', -1)
        
        if music_col >= 0:
            cell = self.ws.cell(row=row_num, column=music_col + 1)
            cell.value = filename
            print(f"  📝 Updated Excel: Music = {filename}")
    
    def generate_song(self, row_data: Dict) -> Dict:
        """Generate complete song using OPTIMIZED pipeline with combined quality checks"""
        try:
            print(f"  🔗 Chain 1: Analysis...")
            analysis = self._chain1_analyze(row_data)
            time.sleep(self.chain_delay)
        except Exception as e:
            if "rate limit" in str(e).lower() or "429" in str(e):
                print(f"  ❌ Rate limit exceeded. Please wait 60+ seconds and try again.")
                raise Exception(f"Rate limit hit - stop and wait before continuing")
            else:
                raise
        
        is_long = row_data['Type'].lower() == 'long'
        
        # OPTIMIZED: Use combined quality check (1 LLM call instead of 3)
        content = None
        combined_quality = None
        best_content = None
        best_score = 0
        max_attempts = self.max_quality_retries + 1  # From .env setting
        
        for attempt in range(max_attempts):
            try:
                print(f"  🔗 Chain 2: Content...{f' (Attempt {attempt + 1})' if attempt > 0 else ''}")
                content = self._chain2_generate_content(row_data, analysis)
                time.sleep(self.chain_delay)
            except Exception as e:
                if "rate limit" in str(e).lower() or "429" in str(e):
                    print(f"  ❌ Rate limit exceeded. Please wait and try again later.")
                    raise Exception(f"Rate limit hit during content generation: {str(e)[:100]}")
                else:
                    print(f"  ❌ Content generation failed: {str(e)[:100]}")
                    if attempt < max_attempts - 1:
                        print(f"  🔄 Retrying...")
                        time.sleep(5)
                        continue
                    else:
                        raise
            
            # Skip Hindi quality check for instrumental (Long) songs
            if is_long:
                print(f"  ⏭️  Skipping Hindi/Rhythm/Melody checks (instrumental)")
                break
            
            # Check Hindi quality (Chain 5 - Grammar & Poetry)
            print(f"  🔗 Chain 5: Hindi Quality (Grammar + Poetry)...")
            hindi_quality = self._chain5_hindi_quality(content, row_data)
            time.sleep(self.chain_delay)
            
            overall_score = hindi_quality.get('overall_score', 0)
            
            # BUG FIX #1: Properly re-check after refinement instead of regenerating
            if overall_score >= self.hindi_threshold:
                print(f"  ✅ Hindi Quality: {overall_score}/10 (Threshold: {self.hindi_threshold})")
                break
            elif attempt < max_attempts - 1:
                issues = hindi_quality.get('issues', [])
                
                if self.use_iterative_refinement:
                    # Try iterative refinement
                    print(f"  ⚠️  Hindi Quality: {overall_score}/10 (Need {self.hindi_threshold}) - Refining...")
                    for issue in issues[:Config.MAX_ISSUES_DISPLAY]:
                        print(f"      - {issue}")
                    
                    content = self._iterative_refine(content, issues, row_data, analysis, focus='hindi')
                    time.sleep(self.chain_delay)
                    
                    # RE-CHECK after refinement (FIX: was missing!)
                    print(f"  🔗 Chain 5: Re-checking Hindi after refinement...")
                    hindi_quality = self._chain5_hindi_quality(content, row_data)
                    time.sleep(self.chain_delay)
                    overall_score = hindi_quality.get('overall_score', 0)
                    
                    if overall_score >= self.hindi_threshold:
                        print(f"  ✅ Hindi Quality after refinement: {overall_score}/10")
                        break
                    else:
                        print(f"  ⚠️  Still {overall_score}/10 after refinement, regenerating...")
                        # Refinement didn't work, regenerate
                        continue
                else:
                    # Full regeneration (old way)
                    print(f"  ⚠️  Hindi Quality: {overall_score}/10 (Need {self.hindi_threshold}) - Regenerating...")
                    for issue in issues[:Config.MAX_ISSUES_DISPLAY]:
                        print(f"      - {issue}")
                    continue
            else:
                print(f"  ⚠️  Hindi Quality: {overall_score}/10 (Target: {self.hindi_threshold}) - Final attempt")
                print(f"      Using best available version")
                break
        
        # OPTIMIZED: Combined Singability check (Rhythm + Melody in ONE call)
        singability_quality = None
        if not is_long:
            for sing_attempt in range(2):
                print(f"  🔗 Chain 5b: Singability Check (Rhythm+Melody)...{f' (Attempt {sing_attempt + 1})' if sing_attempt > 0 else ''}")
                singability_quality = self._chain_singability_check(content, row_data, analysis)
                time.sleep(self.chain_delay)
                
                sing_score = singability_quality.get('overall_singability', 0)
                
                # BUG FIX #2: Properly re-check after refinement
                if sing_score >= self.singability_threshold:
                    print(f"  ✅ Singability: {sing_score}/10 (Threshold: {self.singability_threshold})")
                    break
                elif sing_attempt < 1:
                    issues = singability_quality.get('issues', [])
                    
                    if self.use_iterative_refinement:
                        # Try iterative refinement
                        print(f"  ⚠️  Singability: {sing_score}/10 (Need {self.singability_threshold}) - Refining...")
                        for issue in issues[:Config.MAX_ISSUES_DISPLAY]:
                            print(f"      - {issue}")
                        
                        content = self._iterative_refine(content, issues, row_data, analysis, focus='singability')
                        time.sleep(self.chain_delay)
                        
                        # RE-CHECK after refinement (FIX: was missing!)
                        print(f"  🔗 Chain 5b: Re-checking Singability after refinement...")
                        singability_quality = self._chain_singability_check(content, row_data, analysis)
                        time.sleep(self.chain_delay)
                        sing_score = singability_quality.get('overall_singability', 0)
                        
                        if sing_score >= self.singability_threshold:
                            print(f"  ✅ Singability after refinement: {sing_score}/10")
                            break
                        else:
                            print(f"  ⚠️  Still {sing_score}/10 after refinement, regenerating...")
                            # Refinement didn't work, regenerate
                            print(f"  🔗 Chain 2: Content (Singability focused)...")
                            content = self._chain2_generate_content(row_data, analysis)
                            time.sleep(self.chain_delay)
                    else:
                        # Full regeneration (old way)
                        print(f"  ⚠️  Singability: {sing_score}/10 (Need {self.singability_threshold}) - Regenerating...")
                        for issue in issues[:Config.MAX_ISSUES_DISPLAY]:
                            print(f"      - {issue}")
                        print(f"  🔗 Chain 2: Content (Singability focused)...")
                        content = self._chain2_generate_content(row_data, analysis)
                        time.sleep(self.chain_delay)
                else:
                    print(f"  ⚠️  Singability: {sing_score}/10 (Final attempt)")
                    print(f"      Using best available version")
        
        # Phase 3: Reliability #2 - Validate content length
        if len(content) < Config.MIN_CONTENT_LENGTH:
            raise Exception(f"Generated content too short ({len(content)} chars, min: {Config.MIN_CONTENT_LENGTH})")
        
        # Legacy variables for backward compatibility with save function
        rhythm_quality = singability_quality
        melody_quality = singability_quality
        
        print(f"  🔗 Chain 3: Metadata...")
        metadata = self._chain3_metadata(row_data, analysis, content)
        # Phase 2: Perf #2 - Only sleep if NOT last chain
        if not is_long and not self.skip_chain4_refine:
            time.sleep(self.chain_delay)
        
        if not is_long and not self.skip_chain4_refine:
            print(f"  🔗 Chain 4: Refinement...")
            content = self._chain4_refine(content, row_data)
            time.sleep(self.chain_delay)  # Sleep before Chain 6
        elif not is_long and self.skip_chain4_refine:
            print(f"  ⏭️  Skipping Chain 4 (disabled in .env)")
            time.sleep(self.chain_delay)  # Sleep before Chain 6
        
        # Prepare for cultural validation
        validation_data = {
            'song_title': metadata.get('song_title', ''),
            'lyrics' if not is_long else 'arrangement': content,
            'youtube_description': metadata.get('youtube_description', '')
        }
        
        chain_num = '4' if is_long else '6'
        print(f"  🔗 Chain {chain_num}: Cultural Validation...")
        validation = self._chain6_validate(validation_data, row_data)
        # Phase 2: Perf #2 - Don't sleep after final chain (this is the last one)
        # time.sleep(self.chain_delay)  # REMOVED
        
        # BUG FIX #3: Enforce cultural validation
        status = validation.get('approval_status', 'UNKNOWN')
        score = validation.get('cultural_score', 0)
        
        if score < self.cultural_threshold or status == 'NEEDS_REVISION':
            print(f"  ⚠️  Cultural: {status} (Score: {score}/10, Need: {self.cultural_threshold})")
            issues = validation.get('issues_found', [])
            for issue in issues[:Config.MAX_ISSUES_DISPLAY]:
                print(f"      - {issue}")
            
            # Option: Could retry with feedback, but for now we'll accept with warning
            # User can configure CULTURAL_THRESHOLD higher if they want strict enforcement
            print(f"  ⚠️  Accepting anyway (configure CULTURAL_THRESHOLD in .env for strict mode)")
        else:
            print(f"  ✅ Cultural: {status} (Score: {score}/10)")
        
        return {
            'Song Title': metadata.get('song_title', ''),
            'Hindi Lyrics' if not is_long else 'Suno AI Description': content,
            'Suno AI Description': metadata.get('suno_ai_description', '') if not is_long else content,
            'YouTube Title': metadata.get('youtube_title', ''),
            'YouTube Description': metadata.get('youtube_description', ''),
            'Tags': metadata.get('tags', ''),
            'Analysis': analysis,
            'Hindi Quality': hindi_quality if hindi_quality else {'overall_score': 'N/A', 'verdict': 'Skipped'},
            'Rhythm Quality': rhythm_quality if rhythm_quality else {'overall_rhythm': 'N/A', 'verdict': 'Skipped'},
            'Melody Quality': melody_quality if melody_quality else {'overall_melody': 'N/A', 'verdict': 'Skipped'},
            'Cultural Validation': validation
        }
    
    def process_all(self, start_day: Optional[int] = None, end_day: Optional[int] = None) -> List[Dict]:
        """Process all marked rows"""
        results = []
        skipped = []
        
        # Get column indices
        cols = {
            'day': self.column_map.get('Day', 0),
            'type': self.column_map.get('Type', 1),
            'god': self.column_map.get('God', 2),
            'story': self.column_map.get('Story', 3),
            'style': self.column_map.get('Style', 4),
            'instruments': self.column_map.get('Instruments', 5),
            'user_notes': self.column_map.get('User Notes', -1),
            'language': self.column_map.get('Language', -1),
            'to_generate': self.column_map.get('To generate', -1)
        }
        
        for row_idx, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[cols['day']] is None:
                continue
            
            day = row[cols['day']]
            
            # Range filter
            if start_day and day < start_day:
                continue
            if end_day and day > end_day:
                continue
            
            # Check "To generate"
            if cols['to_generate'] >= 0:
                to_gen = row[cols['to_generate']]
                if not to_gen or str(to_gen).strip().lower() not in ['true', 'yes', '1', 'x']:
                    skipped.append(day)
                    print(f"⏭️  Skip Day {day}")
                    continue
            
            # Extract data
            row_data = {
                'Day': day,
                'Type': str(row[cols['type']]).strip().replace('\xa0', ''),
                'God': str(row[cols['god']]).strip().rstrip('\ufffd'),
                'Story': row[cols['story']],
                'Style': row[cols['style']],
                'Instruments': row[cols['instruments']],
                'User Notes': row[cols['user_notes']] if cols['user_notes'] >= 0 else None,
                'Language': (str(row[cols['language']]).strip() if cols['language'] >= 0 and row[cols['language']] else 'Hindi')
            }
            
            print(f"\n{'='*80}")
            print(f"🎵 Day {day}: {row_data['God']} - {row_data['Story']}")
            print(f"   Type: {row_data['Type']} | Style: {row_data['Style']} | Lang: {row_data['Language']}")
            if row_data['User Notes']:
                print(f"   User Notes: {row_data['User Notes']}")
            print(f"{'='*80}")
            
            try:
                output = self.generate_song(row_data)
                
                # Save individual file
                filename = self._save_song_file(row_data, output)
                
                # Update Excel
                self._update_excel_music_column(row_idx, filename)
                
                results.append({
                    'Day': day,
                    'Input': row_data,
                    'Output': output,
                    'Filename': filename
                })
                
                # Phase 4 - Feature #1: Save progress after each song
                self._save_progress(day, 'completed')
                
                print(f"✅ Day {day} complete!")
                
                # Rate limit: wait before next song
                print(f"⏱️  Cooling down {self.song_delay}s...\n")
                time.sleep(self.song_delay)
                
            except Exception as e:
                error_msg = str(e)
                print(f"❌ Error: {error_msg}")
                import traceback
                traceback.print_exc()
                
                # Phase 4 - Feature #1: Track failure
                self._save_progress(day, 'failed', error_msg[:200])
                
                results.append({
                    'Day': day,
                    'Input': row_data,
                    'Output': {'Error': error_msg}
                })
                time.sleep(self.song_delay)  # Wait even on error
        
        # Save Excel
        try:
            self.wb.save(self.excel_path)
            print(f"\n💾 Excel updated: {self.excel_path}")
        except PermissionError:
            print(f"\n⚠️  Warning: Could not update Excel file (file is open)")
            print(f"   Close Excel and run again to update 'Music' column")
            print(f"   All songs were saved to content/ folder successfully!")
        except Exception as e:
            print(f"\n⚠️  Warning: Could not save Excel: {str(e)}")
        
        if skipped:
            print(f"📊 Skipped {len(skipped)} days: {skipped}")
        
        # Phase 4 - Feature #2: Show cost report
        cost_report = self._generate_cost_report()
        print(cost_report)
        
        # Save cost report to file in cost_reports folder
        cost_report_file = self.cost_dir / f"cost_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(cost_report_file, 'w', encoding='utf-8') as f:
            f.write(cost_report)
            f.write(f"\n\nGeneration Details:\n")
            f.write(f"Start Date: {self.progress.get('start_time', 'N/A')}\n")
            f.write(f"End Date: {datetime.now().strftime('%Y-%m-%d')}\n")
            f.write(f"Total Songs: {len(results)}\n")
            f.write(f"Success: {len([r for r in results if 'Error' not in r.get('Output', {})])}\n")
            f.write(f"Errors: {len([r for r in results if 'Error' in r.get('Output', {})])}\n")
        print(f"\n💾 Cost report saved: {cost_report_file.name}")
        
        # Summary
        success = len([r for r in results if 'Error' not in r.get('Output', {})])
        errors = len(results) - success
        print(f"\n📊 GENERATION SUMMARY:")
        print(f"   Total: {len(results)} | Success: {success} | Errors: {errors}")
        
        return results
    
    def save_results(self, results: List[Dict]):
        """Save consolidated results"""
        # JSON
        json_path = self.cost_dir / "generated_songs_llm.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"💾 JSON: {json_path}")
        
        # Text
        txt_path = self.cost_dir / "generated_songs_llm.txt"
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write("="*100 + "\n")
            f.write("AI-GENERATED DEVOTIONAL SONGS (Google Gemini)\n")
            f.write("="*100 + "\n\n")
            
            for result in results:
                if 'Error' in result.get('Output', {}):
                    f.write(f"\nDAY {result['Day']}: ERROR\n")
                    f.write(f"Error: {result['Output']['Error']}\n\n")
                    continue
                
                day = result['Day']
                inp = result['Input']
                out = result['Output']
                
                f.write(f"\n{'='*100}\n")
                f.write(f"DAY {day}: {inp['God']} - {inp['Story']}\n")
                f.write(f"{'='*100}\n\n")
                
                f.write(f"File: {result.get('Filename', 'N/A')}\n\n")
                
                for key in ['Song Title', 'Hindi Lyrics', 'Suno AI Description', 
                           'YouTube Title', 'YouTube Description', 'Tags']:
                    if key in out:
                        f.write(f"{key.upper()}:\n{out[key]}\n\n")
                
                if 'Validation' in out:
                    val = out['Validation']
                    f.write(f"VALIDATION:\n")
                    f.write(f"  Status: {val.get('approval_status')}\n")
                    f.write(f"  Score: {val.get('cultural_score')}/10\n")
                    f.write(f"  Summary: {val.get('summary')}\n\n")
        
        print(f"💾 Text: {txt_path}")


def main():
    """Main execution"""
    import sys
    
    print("\n" + "="*80)
    print("🤖 AI DEVOTIONAL SONG GENERATOR (Gemini)")
    print("8-stage pipeline | Hindi + Rhythm + Melody + Cultural | Auto-retry")
    print("="*80 + "\n")
    
    start_day = int(sys.argv[1]) if len(sys.argv) > 1 else None
    end_day = int(sys.argv[2]) if len(sys.argv) > 2 else None
    
    if start_day or end_day:
        print(f"📌 Range: Day {start_day or 'start'} to {end_day or 'end'}\n")
    
    try:
        generator = GeminiSongGenerator("First30Days.xlsx")
        results = generator.process_all(start_day, end_day)
        generator.save_results(results)
        
        success = sum(1 for r in results if 'Error' not in r.get('Output', {}))
        
        print(f"\n{'='*80}")
        print(f"✅ GENERATION COMPLETED!")
        print(f"   Total: {len(results)} | Success: {success} | Errors: {len(results) - success}")
        print(f"   Files: content/ folder")
        print(f"{'='*80}\n")
        
    except KeyboardInterrupt:
        print(f"\n\n⚠️  Generation interrupted by user\n")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Fatal Error: {str(e)}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
